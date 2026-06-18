from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from typing import Optional
import io
import logging
from collections import defaultdict
from ..database import get_db

logger = logging.getLogger(__name__)
from ..models import Import, Tarefa, ProgramacaoSemanal, Semana, SubTarefa
from ..schemas import ImportOut, ImportResultado
from ..parsers.xlsx_parser import parse_xlsx
from ..parsers.xer_parser import parse_xer

router = APIRouter(prefix="/imports", tags=["imports"])


@router.get("/", response_model=list[ImportOut])
def listar_imports(db: Session = Depends(get_db)):
    return db.query(Import).order_by(Import.importado_em.desc()).all()


@router.delete("/")
def limpar_historico(db: Session = Depends(get_db)):
    db.query(Import).delete()
    db.commit()
    return {"ok": True}


@router.post("/xlsx", response_model=ImportResultado)
async def importar_xlsx(
    file: UploadFile = File(...),
    semana: str = Form(...),
    aba: Optional[str] = Form(None),
    usuario: Optional[str] = Form("sistema"),
    disciplinas: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    """
    Importa um arquivo XLSX, aplica filtro QCRON e persiste as tarefas.
    """
    semana_obj = db.query(Semana).filter(Semana.codigo == semana).first()
    if not semana_obj:
        raise HTTPException(status_code=404, detail=f"Semana '{semana}' não encontrada. Crie-a primeiro.")

    conteudo = await file.read()
    try:
        tarefas_raw = parse_xlsx(io.BytesIO(conteudo), aba_preferida=aba)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Erro ao ler XLSX: {str(e)}")

    disciplinas_filtro = [d.strip() for d in disciplinas.split(',')] if disciplinas else None
    resultado = _persistir_tarefas(tarefas_raw, semana_obj, db, disciplinas_filtro=disciplinas_filtro)

    registro = Import(
        tipo="xlsx",
        semana_ref=semana,
        arquivo_original=file.filename,
        usuario=usuario,
        status="ok",
    )
    db.add(registro)
    db.commit()
    db.refresh(registro)

    # Recalcula painel de avanço físico automaticamente.
    # Falha aqui NÃO invalida o import — mas precisa aparecer no log
    # para que o planejador saiba que o Painel pode estar desatualizado.
    try:
        from ..services.painel_calc import calcular_painel
        calcular_painel(semana_obj, db)
    except Exception:
        logger.exception("Falha ao recalcular painel após import da semana %s", semana)

    return ImportResultado(
        importacao_id=registro.id,
        semana=semana,
        **resultado,
    )


@router.post("/xer", response_model=ImportResultado)
async def importar_xer(
    file: UploadFile = File(...),
    semana: str = Form(...),
    usuario: Optional[str] = Form("sistema"),
    disciplinas: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    semana_obj = db.query(Semana).filter(Semana.codigo == semana).first()
    if not semana_obj:
        raise HTTPException(status_code=404, detail=f"Semana '{semana}' não encontrada.")

    conteudo = await file.read()
    try:
        tarefas_raw = parse_xer(conteudo.decode("latin-1"))
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Erro ao ler XER: {str(e)}")

    disciplinas_filtro = [d.strip() for d in disciplinas.split(',')] if disciplinas else None
    resultado = _persistir_tarefas(tarefas_raw, semana_obj, db, disciplinas_filtro=disciplinas_filtro)

    registro = Import(
        tipo="xer",
        semana_ref=semana,
        arquivo_original=file.filename,
        usuario=usuario,
        status="ok",
    )
    db.add(registro)
    db.commit()
    db.refresh(registro)

    # Recalcula painel de avanço físico automaticamente.
    # Falha aqui NÃO invalida o import — mas precisa aparecer no log
    # para que o planejador saiba que o Painel pode estar desatualizado.
    try:
        from ..services.painel_calc import calcular_painel
        calcular_painel(semana_obj, db)
    except Exception:
        logger.exception("Falha ao recalcular painel após import da semana %s", semana)

    return ImportResultado(
        importacao_id=registro.id,
        semana=semana,
        **resultado,
    )


@router.post("/semanas")
async def importar_semanas_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Lê a aba 'Semanas' de um arquivo XLSX e cria/atualiza semanas no banco.
    A aba deve ter: coluna B = data, coluna C = código (ex: S_35).
    """
    import openpyxl
    from datetime import datetime as dt

    conteudo = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Erro ao abrir arquivo: {e}")

    if "Semanas" not in wb.sheetnames:
        raise HTTPException(status_code=422, detail="Aba 'Semanas' não encontrada no arquivo.")

    ws = wb["Semanas"]

    # Agrupa datas por código de semana
    semanas_map = defaultdict(list)
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 3:
            continue
        data_val, codigo = row[1], row[2]
        if not data_val or not codigo:
            continue
        if isinstance(data_val, dt):
            data_val = data_val.date()
        semanas_map[str(codigo).strip()].append(data_val)

    if not semanas_map:
        raise HTTPException(status_code=422, detail="Nenhuma semana encontrada na aba 'Semanas'.")

    criadas, atualizadas = 0, 0
    for codigo, datas in sorted(semanas_map.items()):
        data_inicio = min(datas)
        data_fim = max(datas)
        existing = db.query(Semana).filter(Semana.codigo == codigo).first()
        if existing:
            existing.data_inicio = data_inicio
            existing.data_fim = data_fim
            atualizadas += 1
        else:
            db.add(Semana(codigo=codigo, data_inicio=data_inicio, data_fim=data_fim))
            criadas += 1

    db.commit()
    return {"criadas": criadas, "atualizadas": atualizadas, "total": criadas + atualizadas}


def _persistir_tarefas(tarefas_raw: list[dict], semana_obj: Semana, db: Session, disciplinas_filtro: list[str] = None) -> dict:
    """
    Upsert de tarefas e criação de programacao_semanal com filtro QCRON.

    Na reimportação:
    - Preserva estado QPROG (no_qprog, datas, observações) das atividades já programadas
    - Auto-QREAL: atividades que estavam no QPROG e chegaram a 100% de execução são
      marcadas automaticamente como qreal_concluida=True
    - Preserva sub-tarefas cadastradas manualmente
    """
    from ..parsers.xlsx_parser import calcular_qcron

    tarefas_novas = 0
    tarefas_atualizadas = 0
    tarefas_persistidas = []

    for t in tarefas_raw:
        activity_id = t.get("activity_id")
        if not activity_id:
            continue

        existing = db.query(Tarefa).filter(Tarefa.activity_id == activity_id).first()
        if existing:
            for campo in ["nome", "disciplina", "supervisor", "encarregado",
                          "area_unidade", "wbs_codigo", "wbs_path", "duracao", "inicio_lb", "termino_lb"]:
                if t.get(campo) is not None:
                    setattr(existing, campo, t[campo])
            # Sempre atualiza datas reais do cronograma
            if t.get("inicio_prog") is not None:
                existing.inicio_atual = t["inicio_prog"]
            if t.get("termino_prog") is not None:
                existing.termino_atual = t["termino_prog"]
            # Atualiza % físico e peso SMO (usados pelo painel de avanço)
            if t.get("pct_avanco") is not None:
                existing.pct_avanco = t["pct_avanco"]
            if t.get("unid_orcadas_smo") is not None:
                existing.unid_orcadas_smo = t["unid_orcadas_smo"]
            db.flush()
            tarefa = existing
            tarefas_atualizadas += 1
        else:
            tarefa = Tarefa(
                activity_id=activity_id,
                nome=t.get("nome", activity_id),
                disciplina=t.get("disciplina"),
                supervisor=t.get("supervisor"),
                encarregado=t.get("encarregado"),
                area_unidade=t.get("area_unidade"),
                wbs_codigo=t.get("wbs_codigo"),
                wbs_path=t.get("wbs_path"),
                duracao=t.get("duracao"),
                inicio_lb=t.get("inicio_lb"),
                termino_lb=t.get("termino_lb"),
                inicio_atual=t.get("inicio_prog"),
                termino_atual=t.get("termino_prog"),
                pct_avanco=t.get("pct_avanco", 0.0),
                unid_orcadas_smo=t.get("unid_orcadas_smo"),
            )
            db.add(tarefa)
            db.flush()
            tarefas_novas += 1

        tarefas_persistidas.append((tarefa, t))

    db.flush()

    # ── Salva estado anterior antes de deletar ──────────────────────────────
    # qprog_anteriores: tarefa_id -> dict com dados do QPROG anterior
    # sub_tarefas_anteriores: tarefa_id -> lista de dicts das sub-tarefas
    qprog_anteriores: dict[int, dict] = {}
    sub_tarefas_anteriores: dict[int, list] = {}

    progs_anteriores = db.query(ProgramacaoSemanal).filter(
        ProgramacaoSemanal.semana == semana_obj.codigo
    ).all()

    # adiantadas_anteriores: tarefa_id -> dict com dados do adiantamento
    adiantadas_anteriores: dict[int, dict] = {}

    for p in progs_anteriores:
        if p.no_qprog:
            qprog_anteriores[p.tarefa_id] = {
                "inicio_qprog": p.inicio_qprog,
                "termino_qprog": p.termino_qprog,
                "observacoes": p.observacoes,
                "qreal_concluida": p.qreal_concluida,
                "pct_qreal": p.pct_qreal,
            }
        if p.adiantada:
            adiantadas_anteriores[p.tarefa_id] = {
                "semana_original": p.semana_original,
                "no_qprog": p.no_qprog,
                "inicio_qprog": p.inicio_qprog,
                "termino_qprog": p.termino_qprog,
                "observacoes": p.observacoes,
                "qreal_concluida": p.qreal_concluida,
                "pct_qreal": p.pct_qreal,
            }
        # Preserva sub-tarefas de QUALQUER programação (QPROG ou não)
        subs = db.query(SubTarefa).filter(SubTarefa.programacao_id == p.id).all()
        if subs:
            sub_tarefas_anteriores[p.tarefa_id] = [
                {
                    "descricao": s.descricao,
                    "status": s.status,
                    "inicio_qprog": s.inicio_qprog,
                    "termino_qprog": s.termino_qprog,
                }
                for s in subs
            ]

    # ── Aplica filtro QCRON ─────────────────────────────────────────────────
    qcron_items = calcular_qcron(
        tarefas_persistidas,
        semana_obj.data_inicio,
        semana_obj.data_fim,
    )

    if disciplinas_filtro:
        qcron_items = [(tar, t) for tar, t in qcron_items if tar.disciplina in disciplinas_filtro]

    # ── Auto-QREAL ──────────────────────────────────────────────────────────
    # Atividades que estavam no QPROG e chegaram a 100% (excluídas do novo QCRON)
    qcron_tarefa_ids = {tarefa.id for tarefa, _ in qcron_items}
    auto_qreal_items: list[tuple] = []

    for tarefa, t in tarefas_persistidas:
        # Estava no QPROG E não está no novo QCRON
        if tarefa.id in qprog_anteriores and tarefa.id not in qcron_tarefa_ids:
            pct_exec = t.get("pct_executado")
            pct_prev = t.get("pct_avanco", 0.0) or 0.0
            pct = pct_exec if pct_exec is not None else pct_prev
            if pct >= 100:
                auto_qreal_items.append((tarefa, t))

    # ── Apaga e recria ──────────────────────────────────────────────────────
    # Deleta sub-tarefas explicitamente antes das programações.
    # `delete(synchronize_session=False)` não dispara o cascade ORM, então
    # se confiássemos só na cascade, as sub-tarefas ficariam órfãs e
    # duplicariam quando recriássemos via `sub_tarefas_anteriores`.
    prog_ids_antigos = [p.id for p in progs_anteriores]
    if prog_ids_antigos:
        db.query(SubTarefa).filter(
            SubTarefa.programacao_id.in_(prog_ids_antigos)
        ).delete(synchronize_session=False)
    db.query(ProgramacaoSemanal).filter(
        ProgramacaoSemanal.semana == semana_obj.codigo,
    ).delete(synchronize_session=False)
    db.flush()

    def _criar_prog(tarefa, t, qprog_prev: dict | None = None, force_qreal: bool = False, adiantada: bool = False, semana_original: str | None = None):
        """Cria um ProgramacaoSemanal restaurando estado QPROG/QREAL anterior."""
        qp = qprog_prev or {}
        prog = ProgramacaoSemanal(
            semana=semana_obj.codigo,
            tarefa_id=tarefa.id,
            inicio_prog=t.get("inicio_prog"),
            termino_prog=t.get("termino_prog"),
            status_atividade=t.get("status_atividade"),
            pct_avanco=t.get("pct_avanco", 0.0),
            pct_executado=t.get("pct_executado", 0.0),
            # QPROG
            no_qprog=bool(qp),
            inicio_qprog=qp.get("inicio_qprog"),
            termino_qprog=qp.get("termino_qprog"),
            observacoes=qp.get("observacoes"),
            # QREAL
            qreal_concluida=True if force_qreal else qp.get("qreal_concluida", False),
            pct_qreal=100.0 if force_qreal else qp.get("pct_qreal", 0.0),
            # Adiantamento
            adiantada=adiantada,
            semana_original=semana_original,
        )
        db.add(prog)
        db.flush()
        for sub_data in sub_tarefas_anteriores.get(tarefa.id, []):
            db.add(SubTarefa(programacao_id=prog.id, **sub_data))
        return prog

    # Atividades do QCRON normal
    for tarefa, t in qcron_items:
        qprog_prev = qprog_anteriores.get(tarefa.id)
        _criar_prog(tarefa, t, qprog_prev=qprog_prev)

    # Atividades auto-QREAL (estavam no QPROG, chegaram a 100%)
    for tarefa, t in auto_qreal_items:
        qprog_prev = qprog_anteriores[tarefa.id]
        _criar_prog(tarefa, t, qprog_prev=qprog_prev, force_qreal=True)

    # ── Reinsere atividades adiantadas que ainda não concluíram ────────────
    # Condições para reinserir: estava adiantada + não está no novo QCRON + não virou auto-QREAL
    auto_qreal_ids = {tarefa.id for tarefa, _ in auto_qreal_items}
    adiantadas_reinseridas = 0
    for tarefa_id, adiantada_data in adiantadas_anteriores.items():
        if tarefa_id in qcron_tarefa_ids:
            continue  # Já entrou no QCRON natural — não reinserir como adiantada
        if tarefa_id in auto_qreal_ids:
            continue  # Virou auto-QREAL — não reinserir
        # Busca os dados atuais da tarefa no arquivo P6
        tarefa_prog = next(((tar, t) for tar, t in tarefas_persistidas if tar.id == tarefa_id), None)
        if not tarefa_prog:
            continue
        tarefa, t = tarefa_prog
        pct_exec = t.get("pct_executado")
        pct_prev = t.get("pct_avanco", 0.0) or 0.0
        pct = pct_exec if pct_exec is not None else pct_prev
        if pct >= 100:
            continue  # Concluída, não reinserir
        _criar_prog(
            tarefa, t,
            qprog_prev=adiantada_data if adiantada_data.get("no_qprog") else None,
            adiantada=True,
            semana_original=adiantada_data.get("semana_original"),
        )
        adiantadas_reinseridas += 1

    db.commit()

    from ..schemas import TarefaOut
    tarefas_out = [TarefaOut.model_validate(t) for t, _ in tarefas_persistidas]

    return {
        "tarefas_encontradas": len(tarefas_raw),
        "tarefas_novas": tarefas_novas,
        "tarefas_atualizadas": tarefas_atualizadas,
        "qcron_count": len(qcron_items) + len(auto_qreal_items) + adiantadas_reinseridas,
        "auto_qreal_count": len(auto_qreal_items),
        "detalhes": tarefas_out,
    }
