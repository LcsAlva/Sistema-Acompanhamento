"""Serviço de exportação Excel — centralizador de todas as exportações futuras.

Responsabilidade única: gerar workbooks openpyxl a partir de dados já
calculados pelo bm_service.  NÃO recalcula regras financeiras — consome
apenas a saída de montar_bm_completo() e get_curva_s_consolidada().

════════════════════════════════════════════════════════════════════════════════
INVARIANTES
════════════════════════════════════════════════════════════════════════════════
  • Nenhuma query ao banco aqui — recebe dicts prontos.
  • Nenhuma regra financeira — não recalcula nada.
  • Retorna bytes (BytesIO.getvalue()) — agnóstico ao transporte HTTP.
"""
from __future__ import annotations

import hashlib
import io
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ── Paleta corporativa ────────────────────────────────────────────────────────

_AZUL_ESCURO  = "1B3A6B"   # cabeçalho principal — azul Petrobras
_AZUL_MEDIO   = "2E75B6"   # subseção / aba
_AZUL_CLARO   = "BDD7EE"   # zebra par
_BRANCO       = "FFFFFF"
_CINZA_CLARO  = "F2F2F2"   # zebra ímpar
_CINZA_TITULO = "D6DCE4"   # totais / agrupadores
_VERDE        = "375623"    # fg sobre fundo verde
_VERDE_BG     = "E2EFDA"   # fundo positivo
_VERMELHO     = "9C0006"   # fg sobre fundo vermelho
_VERMELHO_BG  = "FFC7CE"   # fundo negativo
_AMARELO_BG   = "FFEB9C"   # alerta / atenção
_AMARELO_FG   = "9C5700"

# ── Helpers de estilo ─────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = "000000", size: int = 10,
          italic: bool = False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic)


def _border_thin() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _border_medium_bottom() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    med  = Side(style="medium", color="888888")
    return Border(left=thin, right=thin, top=thin, bottom=med)


def _align(horizontal: str = "left", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)


def _header_style(ws, row: int, col: int, value: Any,
                  bg: str = _AZUL_ESCURO, fg: str = _BRANCO,
                  size: int = 10, bold: bool = True) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill   = _fill(bg)
    cell.font   = _font(bold=bold, color=fg, size=size)
    cell.border = _border_thin()
    cell.alignment = _align("center")


def _auto_width(ws, min_width: int = 10, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, max_width))


def _zebra(row: int) -> PatternFill:
    return _fill(_AZUL_CLARO) if row % 2 == 0 else _fill(_CINZA_CLARO)


def _fmt_brl(value: float | None) -> str:
    """Formata número como moeda BRL legível (para células de texto/resumo)."""
    if value is None:
        return "R$ 0,00"
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_pct(value: float | None, decimals: int = 2) -> str:
    """Formata 0–1 como 'XX,XX%' (para células de texto/resumo)."""
    if value is None:
        return "0,00%"
    return f"{value * 100:.{decimals}f}%".replace(".", ",")


# ── Formatos numéricos openpyxl ───────────────────────────────────────────────

FMT_BRL  = 'R$ #,##0.00'
FMT_PCT  = '0.00%'
FMT_INT  = '#,##0'
FMT_SPI  = '0.0000'

MES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}

STATUS_LABEL = {
    "em_previa":    "Em Prévia",
    "em_analise":   "Em Análise",
    "pre_aprovada": "Pré-Aprovada",
    "fechada":      "Fechada",
    "consolidada":  "Consolidada",
}


# ════════════════════════════════════════════════════════════════════════════════
# PONTO DE ENTRADA PÚBLICO
# ════════════════════════════════════════════════════════════════════════════════

def gerar_excel_bm(
    bm_data: dict,
    curva_s: list[dict],
    usuario: str = "sistema",
) -> bytes:
    """Gera o workbook Excel do BM mensal e retorna bytes prontos para download.

    Args:
        bm_data:  Saída de montar_bm_completo() — fonte única de dados.
        curva_s:  Saída de get_curva_s_consolidada() — série histórica EVM.
        usuario:  Identificação de quem gerou (para aba Auditoria).

    Returns:
        Bytes do arquivo .xlsx.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove aba vazia padrão

    ciclo    = bm_data["ciclo"]
    itens    = bm_data.get("itens", [])
    pends    = bm_data.get("pendencias", [])
    bac      = float(bm_data.get("bac", 0.0))
    ano, mes = ciclo["ano"], ciclo["mes"]

    # Ponto da curva S para o mês exportado
    iso_mes  = f"{ano}-{mes:02d}-01"
    ponto_ev = next((p for p in curva_s if p["data"] == iso_mes), None)
    pv_acum  = ponto_ev["pv_acum"] if ponto_ev else 0.0
    ev_acum  = ponto_ev["ev_acum"] if ponto_ev else 0.0
    spi      = (ev_acum / pv_acum) if pv_acum > 0 else 0.0

    _aba_resumo(wb, bm_data, pv_acum, ev_acum, spi, bac)
    _aba_medicao(wb, itens, ciclo, bac)
    _aba_pendencias(wb, pends, ciclo)
    _aba_curva_s(wb, curva_s, bac)  # nome da aba: "Curva S - EVM" (sem / que é inválido no Excel)
    _aba_auditoria(wb, bm_data, usuario, len(itens), len(pends))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════════
# ABA 1 — RESUMO EXECUTIVO
# ════════════════════════════════════════════════════════════════════════════════

def _aba_resumo(
    wb: Workbook,
    bm_data: dict,
    pv_acum: float,
    ev_acum: float,
    spi: float,
    bac: float,
) -> None:
    ws = wb.create_sheet("Resumo Executivo")
    ciclo = bm_data["ciclo"]
    ano, mes = ciclo["ano"], ciclo["mes"]

    competencia    = f"{MES_PT[mes]}/{ano}"
    status_label   = STATUS_LABEL.get(ciclo["status"], ciclo["status"])
    total_previsto = float(bm_data.get("total_valor_previsto", 0.0))
    total_medido   = float(bm_data.get("total_valor_periodo", 0.0))
    desvio_val     = float(bm_data.get("desvio_valor_periodo", 0.0))
    total_acum     = float(bm_data.get("total_valor_acum", 0.0))
    pct_acum       = float(bm_data.get("total_pct_acum", 0.0))
    pct_previsto   = float(bm_data.get("total_pct_previsto", 0.0))
    pct_medido     = float(bm_data.get("total_pct_periodo", 0.0))
    pends          = bm_data.get("pendencias", [])
    qtd_pend       = len(pends)
    val_pend       = sum(float(p.get("valor_gap", 0.0)) for p in pends)

    # ── Título ───────────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    titulo = ws["A1"]
    titulo.value     = "BOLETIM DE MEDIÇÃO — RESUMO EXECUTIVO"
    titulo.fill      = _fill(_AZUL_ESCURO)
    titulo.font      = _font(bold=True, color=_BRANCO, size=14)
    titulo.alignment = _align("center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value     = f"Competência: {competencia}  |  BM: {ciclo['numero_bm']}  |  Status: {status_label}"
    sub.fill      = _fill(_AZUL_MEDIO)
    sub.font      = _font(bold=True, color=_BRANCO, size=11)
    sub.alignment = _align("center")
    ws.row_dimensions[2].height = 22

    # ── Função auxiliar de linha ──────────────────────────────────────────────
    def _linha(row: int, label: str, valor: Any,
               fmt: str = "texto", destaque: bool = False) -> None:
        lbl_cell = ws.cell(row=row, column=1, value=label)
        lbl_cell.font      = _font(bold=True, size=10)
        lbl_cell.fill      = _fill(_CINZA_TITULO) if destaque else _fill(_CINZA_CLARO)
        lbl_cell.border    = _border_thin()
        lbl_cell.alignment = _align("left")

        val_cell = ws.cell(row=row, column=2, value=valor)
        val_cell.border    = _border_thin()
        val_cell.alignment = _align("right")
        val_cell.font      = _font(size=10)

        if fmt == "brl":
            val_cell.number_format = FMT_BRL
        elif fmt == "pct":
            val_cell.number_format = FMT_PCT
        elif fmt == "spi":
            val_cell.number_format = FMT_SPI
        elif fmt == "int":
            val_cell.number_format = FMT_INT

        ws.merge_cells(f"C{row}:D{row}")
        info = ws.cell(row=row, column=3)
        info.fill   = _fill(_CINZA_CLARO) if not destaque else _fill(_CINZA_TITULO)
        info.border = _border_thin()

        return val_cell

    # ── Seção: Identificação ──────────────────────────────────────────────────
    r = 4
    ws.merge_cells(f"A{r}:D{r}")
    sec = ws.cell(row=r, column=1, value="IDENTIFICAÇÃO")
    sec.fill      = _fill(_AZUL_MEDIO)
    sec.font      = _font(bold=True, color=_BRANCO, size=10)
    sec.alignment = _align("center")
    r += 1

    _linha(r, "Competência",  competencia);                              r += 1
    _linha(r, "Número BM",    ciclo["numero_bm"]);                       r += 1
    _linha(r, "Status",       status_label, destaque=True);              r += 1
    fechado_em = ciclo.get("fechado_em") or "—"
    if fechado_em != "—":
        try:
            fechado_em = fechado_em[:10]
        except Exception:
            pass
    _linha(r, "Fechado em",   fechado_em);                               r += 1
    _linha(r, "Fechado por",  ciclo.get("fechado_por") or "—");          r += 1

    # ── Seção: Financeiro ─────────────────────────────────────────────────────
    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    sec2 = ws.cell(row=r, column=1, value="FINANCEIRO")
    sec2.fill      = _fill(_AZUL_MEDIO)
    sec2.font      = _font(bold=True, color=_BRANCO, size=10)
    sec2.alignment = _align("center")
    r += 1

    _linha(r, "BAC Total",             bac,             "brl", destaque=True); r += 1
    _linha(r, "Previsto no período",   total_previsto,  "brl");                r += 1
    _linha(r, "% Previsto no período", pct_previsto,    "pct");                r += 1
    _linha(r, "Medido no período",     total_medido,    "brl");                r += 1
    _linha(r, "% Medido no período",   pct_medido,      "pct");                r += 1

    # Desvio com cor condicional
    desvio_cell = _linha(r, "Desvio do período", desvio_val, "brl", destaque=True)
    if desvio_val >= 0:
        desvio_cell.fill = _fill(_VERDE_BG)
        desvio_cell.font = _font(bold=True, color=_VERDE, size=10)
    else:
        desvio_cell.fill = _fill(_VERMELHO_BG)
        desvio_cell.font = _font(bold=True, color=_VERMELHO, size=10)
    r += 1

    _linha(r, "Acumulado atual (R$)",  total_acum,      "brl");  r += 1
    _linha(r, "Acumulado atual (%)",   pct_acum,        "pct");  r += 1

    # ── Seção: EVM ────────────────────────────────────────────────────────────
    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    sec3 = ws.cell(row=r, column=1, value="INDICADORES EVM")
    sec3.fill      = _fill(_AZUL_MEDIO)
    sec3.font      = _font(bold=True, color=_BRANCO, size=10)
    sec3.alignment = _align("center")
    r += 1

    _linha(r, "PV Acumulado",  pv_acum, "brl"); r += 1
    _linha(r, "EV Acumulado",  ev_acum, "brl"); r += 1

    spi_cell = _linha(r, "SPI", spi, "spi", destaque=True)
    if spi >= 1.0:
        spi_cell.fill = _fill(_VERDE_BG)
        spi_cell.font = _font(bold=True, color=_VERDE, size=10)
    elif spi >= 0.85:
        spi_cell.fill = _fill(_AMARELO_BG)
        spi_cell.font = _font(bold=True, color=_AMARELO_FG, size=10)
    else:
        spi_cell.fill = _fill(_VERMELHO_BG)
        spi_cell.font = _font(bold=True, color=_VERMELHO, size=10)
    r += 1

    # ── Seção: Pendências ─────────────────────────────────────────────────────
    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    sec4 = ws.cell(row=r, column=1, value="PENDÊNCIAS")
    sec4.fill      = _fill(_AZUL_MEDIO)
    sec4.font      = _font(bold=True, color=_BRANCO, size=10)
    sec4.alignment = _align("center")
    r += 1

    _linha(r, "Quantidade de pendências", qtd_pend, "int"); r += 1
    pend_val_cell = _linha(r, "Valor total de pendências", val_pend, "brl", destaque=True)
    if qtd_pend > 0:
        pend_val_cell.fill = _fill(_AMARELO_BG)
        pend_val_cell.font = _font(bold=True, color=_AMARELO_FG, size=10)
    r += 1

    # ── Larguras fixas ───────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.freeze_panes = "A3"


# ════════════════════════════════════════════════════════════════════════════════
# ABA 2 — MEDIÇÃO MENSAL
# ════════════════════════════════════════════════════════════════════════════════

_COLS_MEDICAO = [
    ("Código EAP",   14),
    ("Descrição",    45),
    ("Nível",         7),
    ("Valor Item",   16),
    ("% Previsto",   12),
    ("R$ Previsto",  16),
    ("% Medido",     12),
    ("R$ Medido",    16),
    ("Acum. Ant. %", 12),
    ("Acum. Atual %",12),
    ("Saldo %",      11),
    ("Saldo R$",     16),
    ("Status",       14),
]


def _status_item(pct_acum: float, pct_prev: float) -> str:
    if pct_acum >= 1.0 - 1e-4:
        return "Concluído"
    if pct_acum > pct_prev + 1e-4:
        return "Adiantado"
    if pct_acum < pct_prev - 1e-4:
        return "Atrasado"
    return "Conforme"


def _aba_medicao(
    wb: Workbook,
    itens: list[dict],
    ciclo: dict,
    bac: float,
) -> None:
    ws = wb.create_sheet("Medição Mensal")
    ano, mes = ciclo["ano"], ciclo["mes"]

    # ── Título ───────────────────────────────────────────────────────────────
    n_cols = len(_COLS_MEDICAO)
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws.cell(row=1, column=1,
                value=f"MEDIÇÃO MENSAL — {MES_PT[mes]}/{ano}  |  {ciclo['numero_bm']}")
    t.fill      = _fill(_AZUL_ESCURO)
    t.font      = _font(bold=True, color=_BRANCO, size=12)
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 26

    # ── Cabeçalho ────────────────────────────────────────────────────────────
    for c_idx, (col_name, col_w) in enumerate(_COLS_MEDICAO, start=1):
        _header_style(ws, 2, c_idx, col_name)
        ws.column_dimensions[get_column_letter(c_idx)].width = col_w

    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}2"
    ws.freeze_panes = "A3"

    # ── Dados ────────────────────────────────────────────────────────────────
    row = 3
    for item in itens:
        nivel       = int(item.get("nivel", 1))
        is_folha    = bool(item.get("is_folha", True))
        pct_prev    = float(item.get("pct_previsto", 0.0))
        val_item    = float(item.get("valor", 0.0))
        val_prev    = float(item.get("valor_previsto", 0.0))
        pct_med     = float(item.get("pct_periodo", 0.0))
        val_med     = float(item.get("valor_periodo", 0.0))
        pct_ant     = float(item.get("pct_acum_anterior", 0.0))
        pct_acum    = float(item.get("pct_acumulado", 0.0))
        saldo_pct   = max(0.0, 1.0 - pct_acum)
        saldo_val   = saldo_pct * val_item
        status_str  = _status_item(pct_acum, pct_prev)

        # Indentação na descrição conforme nível
        descricao = "  " * (nivel - 1) + str(item.get("descricao", ""))

        # Cor de fundo — hierarquia + zebra
        if nivel == 1:
            bg = _AZUL_ESCURO
            font_color = _BRANCO
            bold = True
        elif not is_folha:
            bg = _AZUL_MEDIO
            font_color = _BRANCO
            bold = True
        else:
            bg = _AZUL_CLARO if row % 2 == 0 else _CINZA_CLARO
            font_color = "000000"
            bold = False

        valores = [
            item.get("codigo", ""),
            descricao,
            nivel,
            val_item,
            pct_prev,
            val_prev,
            pct_med,
            val_med,
            pct_ant,
            pct_acum,
            saldo_pct,
            saldo_val,
            status_str,
        ]
        formatos = [
            None, None, FMT_INT,
            FMT_BRL, FMT_PCT, FMT_BRL,
            FMT_PCT, FMT_BRL,
            FMT_PCT, FMT_PCT,
            FMT_PCT, FMT_BRL,
            None,
        ]

        for c_idx, (val, fmt) in enumerate(zip(valores, formatos), start=1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.fill      = _fill(bg)
            cell.font      = _font(bold=bold, color=font_color, size=10)
            cell.border    = _border_thin()
            cell.alignment = _align(
                "right" if c_idx >= 4 else "left",
                wrap=(c_idx == 2)
            )
            if fmt:
                cell.number_format = fmt

        # Cor condicional na coluna Status (coluna 13)
        status_cell = ws.cell(row=row, column=13)
        if status_str == "Concluído":
            status_cell.fill = _fill(_VERDE_BG)
            status_cell.font = _font(bold=True, color=_VERDE, size=10)
        elif status_str == "Atrasado":
            status_cell.fill = _fill(_VERMELHO_BG)
            status_cell.font = _font(bold=True, color=_VERMELHO, size=10)
        elif status_str == "Adiantado":
            status_cell.fill = _fill(_AMARELO_BG)
            status_cell.font = _font(bold=True, color=_AMARELO_FG, size=10)

        row += 1

    # ── Linha de totais ───────────────────────────────────────────────────────
    if itens:
        ws.merge_cells(f"A{row}:C{row}")
        total_lbl = ws.cell(row=row, column=1, value="TOTAIS (Nível 1)")
        total_lbl.fill      = _fill(_CINZA_TITULO)
        total_lbl.font      = _font(bold=True, size=10)
        total_lbl.border    = _border_medium_bottom()
        total_lbl.alignment = _align("center")

        nivel1 = [i for i in itens if i.get("nivel") == 1]
        totais = {
            4:  sum(float(i.get("valor", 0))          for i in nivel1),
            5:  sum(float(i.get("pct_previsto", 0))   for i in nivel1) / max(len(nivel1), 1),
            6:  sum(float(i.get("valor_previsto", 0)) for i in nivel1),
            8:  sum(float(i.get("valor_periodo", 0))  for i in nivel1),
            10: sum(float(i.get("pct_acumulado", 0))  for i in nivel1) / max(len(nivel1), 1),
            11: sum(float(i.get("valor_acumulado", 0))for i in nivel1),
        }
        fmts_tot = {4: FMT_BRL, 5: FMT_PCT, 6: FMT_BRL, 8: FMT_BRL,
                    10: FMT_PCT, 11: FMT_BRL}
        for col, val in totais.items():
            c = ws.cell(row=row, column=col, value=val)
            c.fill          = _fill(_CINZA_TITULO)
            c.font          = _font(bold=True, size=10)
            c.border        = _border_medium_bottom()
            c.alignment     = _align("right")
            c.number_format = fmts_tot.get(col, "")


# ════════════════════════════════════════════════════════════════════════════════
# ABA 3 — PENDÊNCIAS
# ════════════════════════════════════════════════════════════════════════════════

_COLS_PEND = [
    ("Código EAP",        14),
    ("Descrição",         40),
    ("Competência Orig.", 16),
    ("Gap %",             10),
    ("Gap R$",            16),
    ("Saldo %",           10),
    ("Saldo R$",          16),
    ("Status",            18),
    ("Redistribuído?",    14),
    ("Competência Dest.", 16),
]


def _aba_pendencias(wb: Workbook, pends: list[dict], ciclo: dict) -> None:
    ws = wb.create_sheet("Pendências")
    ano, mes = ciclo["ano"], ciclo["mes"]
    n_cols = len(_COLS_PEND)

    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws.cell(row=1, column=1,
                value=f"PENDÊNCIAS — {MES_PT[mes]}/{ano}  |  {ciclo['numero_bm']}")
    t.fill      = _fill(_AZUL_ESCURO)
    t.font      = _font(bold=True, color=_BRANCO, size=12)
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 26

    if not pends:
        ws.merge_cells(f"A3:{get_column_letter(n_cols)}3")
        msg = ws.cell(row=3, column=1,
                      value="Sem pendências geradas para esta competência.")
        msg.fill      = _fill(_VERDE_BG)
        msg.font      = _font(italic=True, color=_VERDE, size=11)
        msg.alignment = _align("center")
        for c_idx, (_, w) in enumerate(_COLS_PEND, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = w
        return

    # Cabeçalho
    for c_idx, (col_name, col_w) in enumerate(_COLS_PEND, start=1):
        _header_style(ws, 2, c_idx, col_name)
        ws.column_dimensions[get_column_letter(c_idx)].width = col_w

    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}2"
    ws.freeze_panes = "A3"

    row = 3
    for pend in pends:
        redistribuicoes = pend.get("redistribuicoes", [])
        redistribuido   = bool(redistribuicoes)
        dest_comp       = redistribuicoes[-1]["destino"] if redistribuicoes else "—"
        status_pend     = pend.get("status", "ativa")

        valores = [
            pend.get("eap_codigo", ""),
            pend.get("eap_descricao", ""),
            f"{pend.get('ano_origem', '')}/{str(pend.get('mes_origem', '')).zfill(2)}",
            float(pend.get("pct_gap", 0.0)) / 100.0,
            float(pend.get("valor_gap", 0.0)),
            float(pend.get("pct_saldo", 0.0)) / 100.0,
            float(pend.get("valor_saldo", 0.0)),
            status_pend,
            "Sim" if redistribuido else "Não",
            dest_comp,
        ]
        formatos = [
            None, None, None,
            FMT_PCT, FMT_BRL,
            FMT_PCT, FMT_BRL,
            None, None, None,
        ]

        bg = _AZUL_CLARO if row % 2 == 0 else _CINZA_CLARO
        for c_idx, (val, fmt) in enumerate(zip(valores, formatos), start=1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.fill      = _fill(bg)
            cell.font      = _font(size=10)
            cell.border    = _border_thin()
            cell.alignment = _align("right" if c_idx >= 4 else "left", wrap=(c_idx == 2))
            if fmt:
                cell.number_format = fmt

        # Status colorido
        st_cell = ws.cell(row=row, column=8)
        if status_pend == "ativa":
            st_cell.fill = _fill(_VERMELHO_BG)
            st_cell.font = _font(bold=True, color=_VERMELHO, size=10)
        elif status_pend == "redistribuida_parcial":
            st_cell.fill = _fill(_AMARELO_BG)
            st_cell.font = _font(bold=True, color=_AMARELO_FG, size=10)
        elif status_pend == "redistribuida_total":
            st_cell.fill = _fill(_VERDE_BG)
            st_cell.font = _font(bold=True, color=_VERDE, size=10)

        row += 1

    # Linha de totais
    ws.merge_cells(f"A{row}:D{row}")
    lbl = ws.cell(row=row, column=1, value="TOTAIS")
    lbl.fill      = _fill(_CINZA_TITULO)
    lbl.font      = _font(bold=True, size=10)
    lbl.border    = _border_medium_bottom()
    lbl.alignment = _align("center")

    tot_gap_val = ws.cell(row=row, column=5, value=sum(float(p.get("valor_gap", 0)) for p in pends))
    tot_gap_val.fill           = _fill(_CINZA_TITULO)
    tot_gap_val.font           = _font(bold=True, size=10)
    tot_gap_val.border         = _border_medium_bottom()
    tot_gap_val.number_format  = FMT_BRL
    tot_gap_val.alignment      = _align("right")

    tot_sal_val = ws.cell(row=row, column=7, value=sum(float(p.get("valor_saldo", 0)) for p in pends))
    tot_sal_val.fill           = _fill(_CINZA_TITULO)
    tot_sal_val.font           = _font(bold=True, size=10)
    tot_sal_val.border         = _border_medium_bottom()
    tot_sal_val.number_format  = FMT_BRL
    tot_sal_val.alignment      = _align("right")


# ════════════════════════════════════════════════════════════════════════════════
# ABA 4 — CURVA S / EVM
# ════════════════════════════════════════════════════════════════════════════════

_COLS_CURVA = [
    ("Competência",  14),
    ("PV Período",   16),
    ("EV Período",   16),
    ("PV Acumulado", 16),
    ("EV Acumulado", 16),
    ("SPI",          10),
    ("% PV",         10),
    ("% EV",         10),
    ("Desvio R$",    16),
]


def _aba_curva_s(wb: Workbook, curva_s: list[dict], bac: float) -> None:
    ws = wb.create_sheet("Curva S - EVM")
    n_cols = len(_COLS_CURVA)

    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws.cell(row=1, column=1, value="CURVA S / EARNED VALUE MANAGEMENT (EVM)")
    t.fill      = _fill(_AZUL_ESCURO)
    t.font      = _font(bold=True, color=_BRANCO, size=12)
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 26

    if bac > 0:
        ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
        bac_cell = ws.cell(row=2, column=1, value=f"BAC Total: {_fmt_brl(bac)}")
        bac_cell.fill      = _fill(_AZUL_MEDIO)
        bac_cell.font      = _font(bold=True, color=_BRANCO, size=10)
        bac_cell.alignment = _align("center")
        header_row = 3
    else:
        header_row = 2

    for c_idx, (col_name, col_w) in enumerate(_COLS_CURVA, start=1):
        _header_style(ws, header_row, c_idx, col_name)
        ws.column_dimensions[get_column_letter(c_idx)].width = col_w

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(n_cols)}{header_row}"
    ws.freeze_panes = f"A{header_row + 1}"

    row = header_row + 1
    for ponto in curva_s:
        pv_mes   = float(ponto.get("pv_mes", 0.0))
        ev_mes   = float(ponto.get("ev_mes", 0.0))
        pv_acum  = float(ponto.get("pv_acum", 0.0))
        ev_acum  = float(ponto.get("ev_acum", 0.0))
        spi      = (ev_acum / pv_acum) if pv_acum > 0 else 0.0
        pct_pv   = float(ponto.get("pct_pv", 0.0))
        pct_ev   = float(ponto.get("pct_ev", 0.0))
        desvio   = ev_acum - pv_acum

        valores  = [
            ponto.get("label", ponto.get("data", "")),
            pv_mes, ev_mes, pv_acum, ev_acum, spi, pct_pv, pct_ev, desvio,
        ]
        formatos = [
            None,
            FMT_BRL, FMT_BRL, FMT_BRL, FMT_BRL,
            FMT_SPI,
            FMT_PCT, FMT_PCT,
            FMT_BRL,
        ]

        bg = _AZUL_CLARO if row % 2 == 0 else _CINZA_CLARO
        for c_idx, (val, fmt) in enumerate(zip(valores, formatos), start=1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.fill      = _fill(bg)
            cell.font      = _font(size=10)
            cell.border    = _border_thin()
            cell.alignment = _align("right" if c_idx > 1 else "left")
            if fmt:
                cell.number_format = fmt

        # SPI colorido
        spi_cell = ws.cell(row=row, column=6)
        if spi >= 1.0:
            spi_cell.fill = _fill(_VERDE_BG)
            spi_cell.font = _font(bold=True, color=_VERDE, size=10)
        elif 0 < spi < 0.85:
            spi_cell.fill = _fill(_VERMELHO_BG)
            spi_cell.font = _font(bold=True, color=_VERMELHO, size=10)

        # Desvio colorido
        dev_cell = ws.cell(row=row, column=9)
        if desvio >= 0:
            dev_cell.fill = _fill(_VERDE_BG)
            dev_cell.font = _font(bold=True, color=_VERDE, size=10)
        else:
            dev_cell.fill = _fill(_VERMELHO_BG)
            dev_cell.font = _font(bold=True, color=_VERMELHO, size=10)

        row += 1

    if not curva_s:
        ws.merge_cells(f"A{header_row + 1}:{get_column_letter(n_cols)}{header_row + 1}")
        msg = ws.cell(row=header_row + 1, column=1,
                      value="Nenhum dado de Curva S disponível.")
        msg.fill      = _fill(_CINZA_CLARO)
        msg.font      = _font(italic=True, size=10)
        msg.alignment = _align("center")


# ════════════════════════════════════════════════════════════════════════════════
# ABA 5 — AUDITORIA
# ════════════════════════════════════════════════════════════════════════════════

def _aba_auditoria(
    wb: Workbook,
    bm_data: dict,
    usuario: str,
    qtd_itens: int,
    qtd_pendencias: int,
) -> None:
    ws = wb.create_sheet("Auditoria")
    ciclo = bm_data["ciclo"]

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="AUDITORIA DE EXPORTAÇÃO")
    t.fill      = _fill(_AZUL_ESCURO)
    t.font      = _font(bold=True, color=_BRANCO, size=12)
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 26

    agora      = datetime.now(timezone.utc)
    agora_str  = agora.strftime("%d/%m/%Y %H:%M:%S UTC")
    export_id  = hashlib.sha256(
        f"{ciclo['numero_bm']}:{agora.isoformat()}:{usuario}".encode()
    ).hexdigest()[:16].upper()

    versoes    = bm_data.get("qtd_itens_previstos", 0)

    rows_audit = [
        ("Data de geração",       agora_str),
        ("Usuário",               usuario),
        ("Sistema",               "ETM Engenharia — URFCC / Petrobras"),
        ("BM exportado",          ciclo.get("numero_bm", "—")),
        ("Competência",           f"{ciclo['ano']}/{ciclo['mes']:02d}"),
        ("Status BM",             STATUS_LABEL.get(ciclo["status"], ciclo["status"])),
        ("Versão snapshot (itens previstos)", versoes),
        ("Quantidade itens EAP",  qtd_itens),
        ("Quantidade pendências", qtd_pendencias),
        ("Export ID",             export_id),
    ]

    for r_idx, (label, valor) in enumerate(rows_audit, start=3):
        lbl = ws.cell(row=r_idx, column=1, value=label)
        lbl.fill      = _fill(_CINZA_CLARO)
        lbl.font      = _font(bold=True, size=10)
        lbl.border    = _border_thin()
        lbl.alignment = _align("left")

        val = ws.cell(row=r_idx, column=2, value=valor)
        val.border    = _border_thin()
        val.alignment = _align("left")
        val.font      = _font(size=10)

        ws.merge_cells(f"C{r_idx}:D{r_idx}")
        ws.cell(row=r_idx, column=3).border = _border_thin()

    # Nota de rastreabilidade
    r_nota = len(rows_audit) + 5
    ws.merge_cells(f"A{r_nota}:D{r_nota}")
    nota = ws.cell(row=r_nota, column=1,
                   value="Este documento foi gerado automaticamente pelo Sistema de Medição URFCC. "
                         "Para validade oficial, confrontar com o BM registrado no sistema.")
    nota.fill      = _fill(_AMARELO_BG)
    nota.font      = _font(italic=True, color=_AMARELO_FG, size=9)
    nota.alignment = _align("left", wrap=True)
    ws.row_dimensions[r_nota].height = 30

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
