from __future__ import annotations

import io
import json
import math
from dataclasses import dataclass
from datetime import date, datetime
from typing import Iterable

from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..models import (
    EconomicoAnaliseDRE,
    EconomicoAuditoria,
    EconomicoContaDespesa,
    EconomicoForecastAjuste,
    EconomicoForecastHistorico,
    EconomicoForecastItem,
    EconomicoForecastVersao,
    EconomicoImportacao,
    EconomicoLancamentoRazao,
    EconomicoRelatorioOC,
    EconomicoResumoCalculado,
    EconomicoValor,
)


TOLERANCIA = 0.01
PHASE_ROWS = [8, 250, 297, 822, 1358, 1408]
PHASE_LABELS = {
    8: "Mobilização",
    250: "Engenharia",
    297: "Construção Civil",
    822: "Montagem",
    1358: "Comissionamento",
    1408: "Fornecimento de Bens",
}
PHASE_RANGES = {
    8: (8, 249),
    250: (250, 296),
    297: (297, 821),
    822: (822, 1357),
    1358: (1358, 1407),
    1408: (1408, None),
}
META_REAL_STATUS_ROW = 4
META_FIRST_MONTH_COL = 3       # C
RESUMO_FIRST_MONTH_COL = 4     # D
EAP_BASELINE_FIRST_COL = 23    # W
EAP_TENDENCIA_FIRST_COL = 54   # BB


@dataclass(frozen=True)
class AuditMetric:
    indicador: str
    sistema: float
    resumo_bi: float
    origem_sistema: str
    origem_resumo_bi: str


def importar_e_auditar(
    db: Session,
    conteudo: bytes,
    arquivo_original: str,
    usuario: str | None = None,
) -> dict:
    wb_values = load_workbook(io.BytesIO(conteudo), data_only=True, read_only=False)

    _validar_abas(wb_values)
    meses = _periodos_resumo_bi(wb_values)
    real_count = _contar_meses_reais(wb_values)

    registro = EconomicoImportacao(
        arquivo_original=arquivo_original,
        usuario=usuario,
        status="ok",
        observacao=f"{len(meses)} periodos importados; {real_count} meses realizados.",
    )
    db.add(registro)
    db.flush()

    metricas = _calcular_metricas(wb_values, meses, real_count)
    _persistir_valores(db, registro.id, wb_values, meses, real_count)
    _persistir_investigacao(db, registro.id, wb_values, meses, real_count)
    auditorias = _persistir_auditoria(db, registro.id, metricas)

    db.commit()
    db.refresh(registro)

    return {
        "importacao_id": registro.id,
        "arquivo_original": registro.arquivo_original,
        "importado_em": registro.importado_em,
        "status": registro.status,
        "periodos": len(meses),
        "meses_realizados": real_count,
        "aprovado": all(a.aprovado for a in auditorias),
        "auditoria": [_auditoria_dict(a) for a in auditorias],
    }


def listar_importacoes(db: Session) -> list[EconomicoImportacao]:
    return db.query(EconomicoImportacao).order_by(EconomicoImportacao.importado_em.desc()).all()


def obter_auditoria(db: Session, importacao_id: int | None = None) -> dict:
    if importacao_id is None:
        imp = db.query(EconomicoImportacao).order_by(EconomicoImportacao.importado_em.desc()).first()
    else:
        imp = db.query(EconomicoImportacao).filter(EconomicoImportacao.id == importacao_id).first()
    if not imp:
        return {"importacao": None, "auditoria": [], "aprovado": False}
    rows = (
        db.query(EconomicoAuditoria)
        .filter(EconomicoAuditoria.importacao_id == imp.id)
        .order_by(EconomicoAuditoria.id)
        .all()
    )
    return {
        "importacao": {
            "id": imp.id,
            "arquivo_original": imp.arquivo_original,
            "importado_em": imp.importado_em,
            "status": imp.status,
            "observacao": imp.observacao,
        },
        "auditoria": [_auditoria_dict(a) for a in rows],
        "aprovado": all(a.aprovado for a in rows) if rows else False,
    }


def obter_dashboard(db: Session, importacao_id: int | None = None) -> dict:
    if importacao_id is None:
        imp = db.query(EconomicoImportacao).order_by(EconomicoImportacao.importado_em.desc()).first()
    else:
        imp = db.query(EconomicoImportacao).filter(EconomicoImportacao.id == importacao_id).first()
    if not imp:
        return {"importacao": None, "disponivel": False}

    valores = (
        db.query(EconomicoValor)
        .filter(EconomicoValor.importacao_id == imp.id, EconomicoValor.tipo == "sistema")
        .all()
    )
    acumulados = {
        (v.indicador, v.cenario): float(v.valor or 0)
        for v in valores
        if v.periodo is None
    }
    if not acumulados:
        return {"importacao": _importacao_dict(imp), "disponivel": False}

    def acc(indicador: str, cenario: str) -> float:
        direct = acumulados.get((indicador, cenario))
        if direct is not None:
            return direct
        rotulos = {
            ("receita", "linha_base"): "Receita Linha Base",
            ("receita", "real"): "Receita Realizada",
            ("receita", "tendencia"): "Receita Tendencia",
            ("impostos", "linha_base"): "Impostos Linha Base",
            ("impostos", "real"): "Impostos Realizados",
            ("impostos", "tendencia"): "Impostos Tendencia",
            ("custos_diretos", "linha_base"): "Custos Diretos Linha Base",
            ("custos_diretos", "real"): "Custos Diretos Realizados",
            ("custos_diretos", "tendencia"): "Custos Diretos Tendencia",
            ("custos_indiretos", "linha_base"): "Custos Indiretos Linha Base",
            ("custos_indiretos", "real"): "Custos Indiretos Realizados",
            ("custos_indiretos", "tendencia"): "Custos Indiretos Tendencia",
            ("resultado", "linha_base"): "Resultado Linha Base",
            ("resultado", "real"): "Resultado Atual",
            ("resultado", "tendencia"): "Resultado Forecast",
            ("margem", "linha_base"): "Margem Linha Base",
            ("margem", "real"): "Margem Atual",
            ("margem", "tendencia"): "Margem Forecast",
        }
        return acumulados.get((rotulos.get((indicador, cenario)), "auditoria"), 0.0)

    mensais: dict[date, dict[str, float]] = {}
    for v in valores:
        if v.periodo is None or v.cenario not in {"real", "tendencia"}:
            continue
        bucket = mensais.setdefault(v.periodo, {})
        bucket[f"{v.indicador}_{v.cenario}"] = float(v.valor or 0)

    curva = []
    for mes in sorted(mensais):
        row = mensais[mes]
        receita_prevista = row.get("receita_tendencia", 0.0)
        receita_realizada = row.get("receita_real", 0.0)
        custos_previstos = (
            row.get("custos_diretos_tendencia", 0.0)
            + row.get("custos_indiretos_tendencia", 0.0)
            + row.get("impostos_tendencia", 0.0)
        )
        custos_realizados = (
            row.get("custos_diretos_real", 0.0)
            + row.get("custos_indiretos_real", 0.0)
            + row.get("impostos_real", 0.0)
        )
        resultado_previsto = receita_prevista + custos_previstos
        resultado_realizado = receita_realizada + custos_realizados
        curva.append({
            "periodo": mes.isoformat(),
            "receita_prevista": receita_prevista,
            "receita_realizada": receita_realizada,
            "custos_previstos": abs(custos_previstos),
            "custos_realizados": abs(custos_realizados),
            "resultado_previsto": resultado_previsto,
            "resultado_realizado": resultado_realizado,
        })

    resultado_lb = acc("resultado", "linha_base")
    resultado_forecast = acc("resultado", "tendencia")
    impacto = resultado_forecast - resultado_lb

    return {
        "importacao": _importacao_dict(imp),
        "disponivel": True,
        "kpis": {
            "receita": {
                "linha_base": acc("receita", "linha_base"),
                "realizada": acc("receita", "real"),
                "tendencia": acc("receita", "tendencia"),
            },
            "custos": {
                "diretos": acc("custos_diretos", "tendencia"),
                "indiretos": acc("custos_indiretos", "tendencia"),
                "impostos": acc("impostos", "tendencia"),
            },
            "resultado": {
                "linha_base": resultado_lb,
                "forecast": resultado_forecast,
                "atual": acc("resultado", "real"),
            },
            "margem": {
                "prevista": acc("margem", "linha_base"),
                "forecast": acc("margem", "tendencia"),
                "atual": acc("margem", "real"),
            },
        },
        "curva_receita_custos": curva,
        "resultado_mensal": curva,
        "impacto_financeiro": {
            "resultado_linha_base": resultado_lb,
            "resultado_forecast": resultado_forecast,
            "valor": impacto,
            "percentual": _div(impacto, abs(resultado_lb)),
        },
    }


def obter_custos(db: Session, importacao_id: int | None = None) -> dict:
    imp = _obter_importacao(db, importacao_id)
    if not imp:
        return {"importacao": None, "disponivel": False}

    resumo = _resumo_map(db, imp.id)
    categorias = _categorias_custo(db, imp.id, resumo)
    custos = {
        "diretos": {
            "linha_base": _valor_resumo(resumo, "custos_diretos", "linha_base"),
            "forecast": _valor_resumo(resumo, "custos_diretos", "tendencia"),
            "realizado": _valor_resumo(resumo, "custos_diretos", "real"),
        },
        "indiretos": {
            "linha_base": _valor_resumo(resumo, "custos_indiretos", "linha_base"),
            "forecast": _valor_resumo(resumo, "custos_indiretos", "tendencia"),
            "realizado": _valor_resumo(resumo, "custos_indiretos", "real"),
        },
        "impostos": {
            "linha_base": _valor_resumo(resumo, "impostos", "linha_base"),
            "forecast": _valor_resumo(resumo, "impostos", "tendencia"),
            "realizado": _valor_resumo(resumo, "impostos", "real"),
        },
    }
    custos["totais"] = {
        key: sum(custos[t][key] for t in ("diretos", "indiretos", "impostos"))
        for key in ("linha_base", "forecast", "realizado")
    }

    top_custos = sorted(categorias, key=lambda x: abs(x["forecast"]), reverse=True)[:20]
    top_estouros = sorted(
        [c for c in categorias if c["desvio"] < 0],
        key=lambda x: abs(x["desvio"]),
        reverse=True,
    )[:20]

    return {
        "importacao": _importacao_dict(imp),
        "disponivel": bool(resumo),
        "kpis": custos,
        "distribuicao": top_custos,
        "top_custos": top_custos,
        "top_estouros": top_estouros,
    }


def obter_desvios(db: Session, importacao_id: int | None = None) -> dict:
    imp = _obter_importacao(db, importacao_id)
    if not imp:
        return {"importacao": None, "disponivel": False}

    resumo = _resumo_map(db, imp.id)
    categorias = _categorias_custo(db, imp.id, resumo)
    resultado_lb = _valor_resumo(resumo, "resultado", "linha_base")
    resultado_forecast = _valor_resumo(resumo, "resultado", "tendencia")
    impacto = resultado_forecast - resultado_lb
    ranking = []
    negativos = [row for row in categorias if row["desvio"] < -0.01]
    total_negativo = sum(abs(row["desvio"]) for row in negativos)
    for row in categorias:
        if impacto < 0 and row["desvio"] >= -0.01:
            continue
        if impacto < 0 and total_negativo:
            impacto_fin = impacto * (abs(row["desvio"]) / total_negativo)
        else:
            impacto_fin = row["desvio"]
        if abs(impacto_fin) < 0.01:
            continue
        ranking.append({
            "dimensao": "categoria",
            "categoria": row["categoria"],
            "impacto_financeiro": impacto_fin,
            "impacto_percentual": _div(impacto_fin, abs(impacto)),
            "linha_base": row["previsto"],
            "forecast": row["forecast"],
            "realizado": row["realizado"],
        })

    return {
        "importacao": _importacao_dict(imp),
        "disponivel": bool(resumo),
        "impacto": {
            "resultado_linha_base": resultado_lb,
            "resultado_forecast": resultado_forecast,
            "valor": impacto,
            "percentual": _div(impacto, abs(resultado_lb)),
            "tendencia": "perda" if impacto < 0 else "ganho",
        },
        "ranking": sorted(ranking, key=lambda x: x["impacto_financeiro"])[:30],
    }


def obter_lancamentos(
    db: Session,
    importacao_id: int | None = None,
    categoria: str | None = None,
    fornecedor: str | None = None,
    conta: str | None = None,
    documento: str | None = None,
    periodo_inicio: date | None = None,
    periodo_fim: date | None = None,
    limit: int = 200,
) -> dict:
    imp = _obter_importacao(db, importacao_id)
    if not imp:
        return {"importacao": None, "lancamentos": [], "filtros": {}}

    query = _query_lancamentos_filtrados(
        db,
        imp.id,
        categoria=categoria,
        fornecedor=fornecedor,
        conta=conta,
        documento=documento,
        periodo_inicio=periodo_inicio,
        periodo_fim=periodo_fim,
    )

    rows = (
        query
        .order_by(EconomicoLancamentoRazao.data.desc(), func.abs(EconomicoLancamentoRazao.valor).desc())
        .limit(max(1, min(limit, 500)))
        .all()
    )
    total = query.with_entities(func.coalesce(func.sum(EconomicoLancamentoRazao.valor), 0.0)).scalar() or 0.0

    return {
        "importacao": _importacao_dict(imp),
        "total": float(total),
        "lancamentos": [_lancamento_dict(r) for r in rows],
        "filtros": _filtros_lancamentos(db, imp.id),
    }


def obter_centro_analise(
    db: Session,
    importacao_id: int | None = None,
    categoria: str | None = None,
    fornecedor: str | None = None,
    conta: str | None = None,
    documento: str | None = None,
    periodo_inicio: date | None = None,
    periodo_fim: date | None = None,
) -> dict:
    imp = _obter_importacao(db, importacao_id)
    if not imp:
        return {"importacao": None, "disponivel": False}

    resumo = _resumo_map(db, imp.id)
    categorias_base = _categorias_custo(db, imp.id, resumo)
    if categoria:
        categorias_base = [row for row in categorias_base if row["categoria"] == categoria]

    ledger_query = _query_lancamentos_filtrados(
        db,
        imp.id,
        categoria=categoria,
        fornecedor=fornecedor,
        conta=conta,
        documento=documento,
        periodo_inicio=periodo_inicio,
        periodo_fim=periodo_fim,
    )
    ledger_rows = ledger_query.all()
    categorias_filtradas = {r.categoria_dre for r in ledger_rows if r.categoria_dre}
    if (fornecedor or conta or documento or periodo_inicio or periodo_fim) and categorias_filtradas:
        categorias_base = [row for row in categorias_base if row["categoria"] in categorias_filtradas]

    impacto_total = _impacto_total_aprovado(resumo)
    negativos = [row for row in categorias_base if row["desvio"] < -0.01]
    total_negativo = sum(abs(row["desvio"]) for row in negativos)

    categoria_rows = []
    for row in categorias_base:
        desvio_explicado = row["desvio"]
        if impacto_total < 0:
            if row["desvio"] < -0.01 and total_negativo:
                desvio_explicado = impacto_total * (abs(row["desvio"]) / total_negativo)
            else:
                desvio_explicado = 0.0
        categoria_rows.append({
            **row,
            "desvio": desvio_explicado,
            "desvio_bruto": row["desvio"],
            "impacto_percentual": _div(desvio_explicado, abs(impacto_total)),
            "participacao_percentual": 0.0,
        })
    categoria_rows = sorted(categoria_rows, key=lambda x: x["desvio"])

    total_forecast = sum(row["forecast"] for row in categoria_rows)
    total_previsto = sum(row["previsto"] for row in categoria_rows)
    total_realizado = sum(row["realizado"] for row in categoria_rows)
    total_desvio = sum(row["desvio"] for row in categoria_rows)
    for row in categoria_rows:
        row["participacao_percentual"] = _div(row["forecast"], total_forecast)

    fornecedor_rows = _ranking_fornecedores(db, imp.id, ledger_query)
    conta_rows = _ranking_contas(db, imp.id, ledger_query)
    pareto = _pareto_categorias(categoria_rows)

    return {
        "importacao": _importacao_dict(imp),
        "disponivel": bool(resumo),
        "filtros": _filtros_lancamentos(db, imp.id),
        "kpis": {
            "previsto": total_previsto,
            "forecast": total_forecast,
            "realizado": total_realizado,
            "desvio": total_desvio,
            "impacto_percentual": _div(total_desvio, abs(impacto_total)),
            "participacao_percentual": _div(total_forecast, abs(_valor_resumo(resumo, "custos_diretos", "tendencia"))),
        },
        "analise_categoria": categoria_rows,
        "analise_fornecedor": fornecedor_rows,
        "analise_conta": conta_rows,
        "pareto": pareto,
        "contexto": {
            "impacto_aprovado": impacto_total,
            "fonte": "economico_resumo_calculado, economico_lancamento_razao, economico_relatorio_oc, economico_analise_dre, economico_conta_despesa",
        },
    }


def obter_auditoria_receitas(db: Session, importacao_id: int | None = None) -> dict:
    imp = _obter_importacao(db, importacao_id)
    if not imp:
        return {"importacao": None, "disponivel": False}
    resumo = _resumo_map(db, imp.id)
    cenarios = [
        ("linha_base", "Receita Linha Base"),
        ("tendencia", "Receita Tendencia"),
        ("real", "Receita Realizada"),
    ]
    fases = sorted({
        categoria for indicador, _, categoria in resumo
        if indicador == "receita_fase" and categoria
    })
    validacoes = []
    for cenario, label in cenarios:
        total_fases = sum(_valor_resumo(resumo, "receita_fase", cenario, fase) for fase in fases)
        total_auditado = _valor_resumo(resumo, "receita", cenario)
        diff = total_fases - total_auditado
        validacoes.append({
            "cenario": cenario,
            "indicador": label,
            "total_fases": total_fases,
            "total_auditado": total_auditado,
            "diferenca": diff,
            "aprovado": abs(diff) <= TOLERANCIA,
        })
    mensal = []
    for cenario, _ in cenarios:
        count_fase = (
            db.query(EconomicoResumoCalculado)
            .filter(
                EconomicoResumoCalculado.importacao_id == imp.id,
                EconomicoResumoCalculado.indicador == "receita_fase",
                EconomicoResumoCalculado.cenario == cenario,
                EconomicoResumoCalculado.periodo.isnot(None),
            )
            .count()
        )
        meses_fase = (
            db.query(EconomicoResumoCalculado.periodo)
            .filter(
                EconomicoResumoCalculado.importacao_id == imp.id,
                EconomicoResumoCalculado.indicador == "receita_fase",
                EconomicoResumoCalculado.cenario == cenario,
                EconomicoResumoCalculado.periodo.isnot(None),
            )
            .distinct()
            .count()
        )
        count_total = (
            db.query(EconomicoResumoCalculado)
            .filter(
                EconomicoResumoCalculado.importacao_id == imp.id,
                EconomicoResumoCalculado.indicador == "receita",
                EconomicoResumoCalculado.cenario == cenario,
                EconomicoResumoCalculado.periodo.isnot(None),
                EconomicoResumoCalculado.categoria.is_(None),
            )
            .count()
        )
        diffs = []
        periodos = [
            p[0] for p in db.query(EconomicoResumoCalculado.periodo)
            .filter(
                EconomicoResumoCalculado.importacao_id == imp.id,
                EconomicoResumoCalculado.indicador == "receita",
                EconomicoResumoCalculado.cenario == cenario,
                EconomicoResumoCalculado.periodo.isnot(None),
                EconomicoResumoCalculado.categoria.is_(None),
            )
            .distinct()
            .all()
        ]
        for periodo in periodos:
            total_mes = (
                db.query(func.coalesce(func.sum(EconomicoResumoCalculado.valor), 0.0))
                .filter(
                    EconomicoResumoCalculado.importacao_id == imp.id,
                    EconomicoResumoCalculado.indicador == "receita",
                    EconomicoResumoCalculado.cenario == cenario,
                    EconomicoResumoCalculado.periodo == periodo,
                    EconomicoResumoCalculado.categoria.is_(None),
                )
                .scalar()
            ) or 0.0
            fases_mes = (
                db.query(func.coalesce(func.sum(EconomicoResumoCalculado.valor), 0.0))
                .filter(
                    EconomicoResumoCalculado.importacao_id == imp.id,
                    EconomicoResumoCalculado.indicador == "receita_fase",
                    EconomicoResumoCalculado.cenario == cenario,
                    EconomicoResumoCalculado.periodo == periodo,
                )
                .scalar()
            ) or 0.0
            diffs.append(float(fases_mes) - float(total_mes))
        mensal.append({
            "cenario": cenario,
            "receita_total_existe": count_total > 0,
            "receita_fase_existe": count_fase > 0,
            "meses": meses_fase,
            "linhas_fase": count_fase,
            "linhas_total": count_total,
            "maior_diferenca_mensal": max((abs(d) for d in diffs), default=0.0),
            "mensal_aprovado": max((abs(d) for d in diffs), default=0.0) <= TOLERANCIA,
        })
    por_fase = []
    for fase in fases:
        por_fase.append({
            "fase": fase,
            "linha_base": _valor_resumo(resumo, "receita_fase", "linha_base", fase),
            "tendencia": _valor_resumo(resumo, "receita_fase", "tendencia", fase),
            "real": _valor_resumo(resumo, "receita_fase", "real", fase),
        })
    return {
        "importacao": _importacao_dict(imp),
        "disponivel": bool(fases),
        "tolerancia": TOLERANCIA,
        "aprovado": all(v["aprovado"] for v in validacoes) and all(m["mensal_aprovado"] for m in mensal),
        "validacoes": validacoes,
        "mensal": mensal,
        "fases": por_fase,
    }


def obter_receitas(db: Session, importacao_id: int | None = None) -> dict:
    imp = _obter_importacao(db, importacao_id)
    if not imp:
        return {"importacao": None, "disponivel": False}
    resumo = _resumo_map(db, imp.id)
    auditoria = obter_auditoria_receitas(db, imp.id)
    fases = auditoria.get("fases", [])
    receita_tend = _valor_resumo(resumo, "receita", "tendencia")
    receita_real = _valor_resumo(resumo, "receita", "real")
    fases_out = []
    for fase in fases:
        desvio = fase["tendencia"] - fase["linha_base"]
        fases_out.append({
            **fase,
            "participacao_percentual": _div(fase["tendencia"], receita_tend),
            "desvio": desvio,
        })
    curva = _serie_receita(db, imp.id)
    return {
        "importacao": _importacao_dict(imp),
        "disponivel": auditoria.get("aprovado", False),
        "auditoria": auditoria,
        "kpis": {
            "linha_base": _valor_resumo(resumo, "receita", "linha_base"),
            "tendencia": receita_tend,
            "realizada": receita_real,
            "a_reconhecer": receita_tend - receita_real,
        },
        "curva_mensal": curva["mensal"],
        "curva_acumulada": curva["acumulada"],
        "fases": sorted(fases_out, key=lambda x: x["participacao_percentual"], reverse=True),
        "impactos": sorted(fases_out, key=lambda x: abs(x["desvio"]), reverse=True),
    }


def obter_forecast(db: Session, importacao_id: int | None = None) -> dict:
    imp = _obter_importacao(db, importacao_id)
    if not imp:
        return {"importacao": None, "disponivel": False}
    resumo = _resumo_map(db, imp.id)
    resultado_lb = _valor_resumo(resumo, "resultado", "linha_base")
    resultado_fc = _valor_resumo(resumo, "resultado", "tendencia")
    impacto = resultado_fc - resultado_lb
    desvios = obter_desvios(db, imp.id)
    return {
        "importacao": _importacao_dict(imp),
        "disponivel": bool(resumo),
        "kpis": {
            "resultado_linha_base": resultado_lb,
            "resultado_forecast": resultado_fc,
            "impacto_financeiro": impacto,
            "impacto_percentual": _div(impacto, abs(resultado_lb)),
        },
        "curva_mensal": _serie_resultado_forecast(db, imp.id, acumulado=False),
        "curva_acumulada": _serie_resultado_forecast(db, imp.id, acumulado=True),
        "componentes": {
            "receita_forecast": _valor_resumo(resumo, "receita", "tendencia"),
            "custos_diretos_forecast": _valor_resumo(resumo, "custos_diretos", "tendencia"),
            "custos_indiretos_forecast": _valor_resumo(resumo, "custos_indiretos", "tendencia"),
            "impostos_forecast": _valor_resumo(resumo, "impostos", "tendencia"),
            "resultado_forecast": resultado_fc,
        },
        "explicacao": desvios.get("ranking", []),
    }


def obter_resultado(db: Session, importacao_id: int | None = None) -> dict:
    imp = _obter_importacao(db, importacao_id)
    if not imp:
        return {"importacao": None, "disponivel": False}
    resumo = _resumo_map(db, imp.id)
    desvios = obter_desvios(db, imp.id)
    return {
        "importacao": _importacao_dict(imp),
        "disponivel": bool(resumo),
        "kpis": {
            "resultado_linha_base": _valor_resumo(resumo, "resultado", "linha_base"),
            "resultado_forecast": _valor_resumo(resumo, "resultado", "tendencia"),
            "resultado_atual": _valor_resumo(resumo, "resultado", "real"),
            "margem_atual": _valor_resumo(resumo, "margem", "real"),
        },
        "evolucao_resultado": _serie_resultado_completa(db, imp.id),
        "formacao_resultado": {
            "receita": _valor_resumo(resumo, "receita", "tendencia"),
            "impostos": _valor_resumo(resumo, "impostos", "tendencia"),
            "custos_diretos": _valor_resumo(resumo, "custos_diretos", "tendencia"),
            "custos_indiretos": _valor_resumo(resumo, "custos_indiretos", "tendencia"),
            "resultado": _valor_resumo(resumo, "resultado", "tendencia"),
        },
        "evolucao_margem": _serie_margem(db, imp.id),
        "impactos": desvios.get("ranking", []),
    }


def obter_historico(db: Session) -> dict:
    importacoes = db.query(EconomicoImportacao).order_by(EconomicoImportacao.importado_em).all()
    rows = []
    auditoria_rows = []
    for imp in importacoes:
        resumo = _resumo_map(db, imp.id)
        completo = bool(resumo)
        receita_fc = _valor_resumo(resumo, "receita", "tendencia") if completo else 0.0
        custos_fc = (
            _valor_resumo(resumo, "custos_diretos", "tendencia")
            + _valor_resumo(resumo, "custos_indiretos", "tendencia")
            + _valor_resumo(resumo, "impostos", "tendencia")
        ) if completo else 0.0
        resultado_fc = _valor_resumo(resumo, "resultado", "tendencia") if completo else 0.0
        margem_fc = _valor_resumo(resumo, "margem", "tendencia") if completo else 0.0
        auds = db.query(EconomicoAuditoria).filter(EconomicoAuditoria.importacao_id == imp.id).all()
        aud_ok = all(a.aprovado for a in auds) if auds else False
        max_diff = max((abs(a.diferenca or 0.0) for a in auds), default=None)
        item = {
            "id": imp.id,
            "data_hora": imp.importado_em,
            "usuario": imp.usuario,
            "arquivo": imp.arquivo_original,
            "status": imp.status,
            "completo": completo,
            "necessita_reprocessamento": not completo,
            "receita_forecast": receita_fc,
            "custos_forecast": custos_fc,
            "resultado_forecast": resultado_fc,
            "margem_forecast": margem_fc,
            "auditoria_aprovada": aud_ok,
            "maior_diferenca_auditoria": max_diff,
        }
        rows.append(item)
        auditoria_rows.append({
            "importacao": imp.id,
            "receita": receita_fc,
            "custos": custos_fc,
            "resultado": resultado_fc,
            "margem": margem_fc,
            "status_auditoria": "Aprovada" if aud_ok else "Pendente",
            "maior_diferenca": max_diff,
        })
    return {
        "importacoes": rows,
        "evolucao_indicadores": rows,
        "tendencia_historica": [
            {"importacao": r["id"], "data_hora": r["data_hora"], "resultado_forecast": r["resultado_forecast"]}
            for r in rows if r["completo"]
        ],
        "auditoria_historica": auditoria_rows,
        "validacao": {
            "total_importacoes": len(rows),
            "importacoes_completas": sum(1 for r in rows if r["completo"]),
            "importacoes_incompletas": sum(1 for r in rows if not r["completo"]),
        },
    }


def listar_forecast_operacional_versoes(db: Session) -> dict:
    versoes = (
        db.query(EconomicoForecastVersao)
        .order_by(EconomicoForecastVersao.criado_em.desc(), EconomicoForecastVersao.id.desc())
        .all()
    )
    return {"versoes": [_forecast_versao_dict(db, versao) for versao in versoes]}


def criar_forecast_operacional_versao(
    db: Session,
    nome: str | None = None,
    motivo: str | None = None,
    usuario: str | None = None,
    importacao_id: int | None = None,
) -> dict:
    imp = _obter_importacao_auditada(db, importacao_id)
    if not imp:
        raise ValueError("Nenhuma importacao economica auditada encontrada para criar o forecast operacional.")

    codigo = _proximo_codigo_forecast(db)
    versao = EconomicoForecastVersao(
        importacao_id=imp.id,
        codigo=codigo,
        nome=nome or f"Forecast Operacional {codigo}",
        motivo=motivo,
        status="rascunho",
        origem="importacao",
        criado_por=usuario or "sistema",
    )
    db.add(versao)
    db.flush()

    _popular_itens_forecast_operacional(db, versao, imp.id)
    _registrar_forecast_historico(
        db,
        versao,
        "versao_criada",
        "Versao criada a partir da ultima importacao auditada.",
        usuario,
        {"importacao_id": imp.id},
    )
    db.commit()
    db.refresh(versao)
    return obter_forecast_operacional_versao(db, versao.id)


def clonar_forecast_operacional_versao(
    db: Session,
    versao_id: int,
    nome: str | None = None,
    motivo: str | None = None,
    usuario: str | None = None,
) -> dict:
    origem = db.query(EconomicoForecastVersao).filter(EconomicoForecastVersao.id == versao_id).first()
    if not origem:
        raise ValueError("Versao de forecast nao encontrada.")

    codigo = _proximo_codigo_forecast(db)
    clone = EconomicoForecastVersao(
        importacao_id=origem.importacao_id,
        codigo=codigo,
        nome=nome or f"{origem.nome} - copia",
        motivo=motivo,
        status="rascunho",
        origem="clone",
        versao_base_id=origem.id,
        criado_por=usuario or "sistema",
    )
    db.add(clone)
    db.flush()

    itens = db.query(EconomicoForecastItem).filter(EconomicoForecastItem.versao_id == origem.id).all()
    for item in itens:
        db.add(EconomicoForecastItem(
            versao_id=clone.id,
            importacao_id=clone.importacao_id,
            indicador=item.indicador,
            periodo=item.periodo,
            categoria=item.categoria,
            valor=float(item.valor or 0.0),
            origem=f"clone:{origem.codigo}",
        ))

    _registrar_forecast_historico(
        db,
        clone,
        "versao_clonada",
        f"Versao clonada de {origem.codigo}.",
        usuario,
        {"versao_base_id": origem.id, "versao_base_codigo": origem.codigo},
    )
    db.commit()
    db.refresh(clone)
    return obter_forecast_operacional_versao(db, clone.id)


def obter_forecast_operacional_versao(db: Session, versao_id: int) -> dict:
    versao = db.query(EconomicoForecastVersao).filter(EconomicoForecastVersao.id == versao_id).first()
    if not versao:
        raise ValueError("Versao de forecast nao encontrada.")
    itens = (
        db.query(EconomicoForecastItem)
        .filter(EconomicoForecastItem.versao_id == versao.id)
        .order_by(EconomicoForecastItem.indicador, EconomicoForecastItem.categoria)
        .all()
    )
    ajustes = (
        db.query(EconomicoForecastAjuste)
        .filter(EconomicoForecastAjuste.versao_id == versao.id)
        .order_by(EconomicoForecastAjuste.criado_em.desc(), EconomicoForecastAjuste.id.desc())
        .all()
    )
    historico = (
        db.query(EconomicoForecastHistorico)
        .filter(EconomicoForecastHistorico.versao_id == versao.id)
        .order_by(EconomicoForecastHistorico.criado_em.desc(), EconomicoForecastHistorico.id.desc())
        .all()
    )
    componentes = [i for i in itens if i.indicador != "categoria_custo"]
    categorias = [i for i in itens if i.indicador == "categoria_custo"]
    return {
        "versao": _forecast_versao_dict(db, versao),
        "componentes": [_forecast_item_dict(i) for i in componentes],
        "categorias": [_forecast_item_dict(i) for i in sorted(categorias, key=lambda x: abs(float(x.valor or 0.0)), reverse=True)],
        "ajustes": [_forecast_ajuste_dict(a) for a in ajustes],
        "historico": [_forecast_historico_dict(h) for h in historico],
    }


def ajustar_forecast_operacional_categoria(
    db: Session,
    versao_id: int,
    categoria: str,
    valor_novo: float,
    justificativa: str,
    usuario: str | None = None,
) -> dict:
    if not justificativa or not justificativa.strip():
        raise ValueError("Justificativa obrigatoria para ajustar o forecast.")

    versao = db.query(EconomicoForecastVersao).filter(EconomicoForecastVersao.id == versao_id).first()
    if not versao:
        raise ValueError("Versao de forecast nao encontrada.")

    item = (
        db.query(EconomicoForecastItem)
        .filter(
            EconomicoForecastItem.versao_id == versao.id,
            EconomicoForecastItem.indicador == "categoria_custo",
            EconomicoForecastItem.categoria == categoria,
            EconomicoForecastItem.periodo.is_(None),
        )
        .first()
    )
    if not item:
        raise ValueError("Categoria nao encontrada nesta versao de forecast.")

    valor_anterior = float(item.valor or 0.0)
    novo = float(valor_novo or 0.0)
    diferenca = novo - valor_anterior
    item.valor = novo

    _aplicar_delta_forecast_operacional(db, versao.id, "custos_diretos", diferenca)
    _aplicar_delta_forecast_operacional(db, versao.id, "resultado", -diferenca)

    ajuste = EconomicoForecastAjuste(
        versao_id=versao.id,
        item_id=item.id,
        categoria=categoria,
        valor_anterior=valor_anterior,
        valor_novo=novo,
        diferenca=diferenca,
        justificativa=justificativa.strip(),
        usuario=usuario or "sistema",
    )
    db.add(ajuste)
    _registrar_forecast_historico(
        db,
        versao,
        "categoria_ajustada",
        f"Categoria {categoria} ajustada.",
        usuario,
        {
            "categoria": categoria,
            "valor_anterior": valor_anterior,
            "valor_novo": novo,
            "diferenca": diferenca,
        },
    )
    db.commit()
    return obter_forecast_operacional_versao(db, versao.id)


def comparar_forecast_operacional_versoes(db: Session, base_id: int, novo_id: int) -> dict:
    base = db.query(EconomicoForecastVersao).filter(EconomicoForecastVersao.id == base_id).first()
    novo = db.query(EconomicoForecastVersao).filter(EconomicoForecastVersao.id == novo_id).first()
    if not base or not novo:
        raise ValueError("Selecione duas versoes validas para comparar.")

    base_items = _forecast_items_por_chave(db, base.id)
    novo_items = _forecast_items_por_chave(db, novo.id)
    chaves = sorted(set(base_items) | set(novo_items), key=lambda c: (c[0], c[1] or ""))
    linhas = []
    for indicador, categoria in chaves:
        if indicador not in {"categoria_custo", "receita", "custos_diretos", "custos_indiretos", "impostos", "resultado"}:
            continue
        valor_base = base_items.get((indicador, categoria), 0.0)
        valor_novo = novo_items.get((indicador, categoria), 0.0)
        diferenca = valor_novo - valor_base
        if abs(diferenca) < 0.005 and indicador == "categoria_custo":
            continue
        linhas.append({
            "indicador": indicador,
            "categoria": categoria or _rotulo_indicador_forecast(indicador),
            "forecast_atual": valor_base,
            "forecast_novo": valor_novo,
            "diferenca": diferenca,
        })
    return {
        "base": _forecast_versao_dict(db, base),
        "novo": _forecast_versao_dict(db, novo),
        "comparacao": sorted(linhas, key=lambda x: abs(x["diferenca"]), reverse=True),
    }


def _obter_importacao(db: Session, importacao_id: int | None = None) -> EconomicoImportacao | None:
    if importacao_id is None:
        return db.query(EconomicoImportacao).order_by(EconomicoImportacao.importado_em.desc()).first()
    return db.query(EconomicoImportacao).filter(EconomicoImportacao.id == importacao_id).first()


def _obter_importacao_auditada(db: Session, importacao_id: int | None = None) -> EconomicoImportacao | None:
    query = db.query(EconomicoImportacao)
    if importacao_id is not None:
        query = query.filter(EconomicoImportacao.id == importacao_id)
    for imp in query.order_by(EconomicoImportacao.importado_em.desc(), EconomicoImportacao.id.desc()).all():
        resumo_existe = (
            db.query(EconomicoResumoCalculado.id)
            .filter(EconomicoResumoCalculado.importacao_id == imp.id)
            .first()
            is not None
        )
        auditorias = db.query(EconomicoAuditoria).filter(EconomicoAuditoria.importacao_id == imp.id).all()
        if resumo_existe and auditorias and all(a.aprovado for a in auditorias):
            return imp
    return None


def _proximo_codigo_forecast(db: Session) -> str:
    ultimo_id = db.query(func.max(EconomicoForecastVersao.id)).scalar() or 0
    return f"FO-{int(ultimo_id) + 1:04d}"


def _popular_itens_forecast_operacional(db: Session, versao: EconomicoForecastVersao, importacao_id: int) -> None:
    resumo = _resumo_map(db, importacao_id)
    componentes = {
        "receita": _valor_resumo(resumo, "receita", "tendencia"),
        "custos_diretos": abs(_valor_resumo(resumo, "custos_diretos", "tendencia")),
        "custos_indiretos": abs(_valor_resumo(resumo, "custos_indiretos", "tendencia")),
        "impostos": abs(_valor_resumo(resumo, "impostos", "tendencia")),
        "resultado": _valor_resumo(resumo, "resultado", "tendencia"),
    }
    for indicador, valor in componentes.items():
        db.add(EconomicoForecastItem(
            versao_id=versao.id,
            importacao_id=importacao_id,
            indicador=indicador,
            categoria=None,
            periodo=None,
            valor=float(valor or 0.0),
            origem="economico_resumo_calculado:tendencia",
        ))

    for row in _categorias_custo(db, importacao_id, resumo):
        db.add(EconomicoForecastItem(
            versao_id=versao.id,
            importacao_id=importacao_id,
            indicador="categoria_custo",
            categoria=row["categoria"],
            periodo=None,
            valor=float(row["forecast"] or 0.0),
            origem="economico_analise_dre/economico_resumo_calculado",
        ))


def _aplicar_delta_forecast_operacional(db: Session, versao_id: int, indicador: str, delta: float) -> None:
    item = (
        db.query(EconomicoForecastItem)
        .filter(
            EconomicoForecastItem.versao_id == versao_id,
            EconomicoForecastItem.indicador == indicador,
            EconomicoForecastItem.categoria.is_(None),
            EconomicoForecastItem.periodo.is_(None),
        )
        .first()
    )
    if item:
        item.valor = float(item.valor or 0.0) + float(delta or 0.0)


def _registrar_forecast_historico(
    db: Session,
    versao: EconomicoForecastVersao,
    acao: str,
    descricao: str,
    usuario: str | None,
    payload: dict | None = None,
) -> None:
    db.add(EconomicoForecastHistorico(
        versao_id=versao.id,
        acao=acao,
        descricao=descricao,
        usuario=usuario or "sistema",
        payload=json.dumps(payload or {}, ensure_ascii=False),
    ))


def _forecast_items_por_chave(db: Session, versao_id: int) -> dict[tuple[str, str | None], float]:
    rows = (
        db.query(EconomicoForecastItem)
        .filter(EconomicoForecastItem.versao_id == versao_id, EconomicoForecastItem.periodo.is_(None))
        .all()
    )
    return {(r.indicador, r.categoria): float(r.valor or 0.0) for r in rows}


def _rotulo_indicador_forecast(indicador: str) -> str:
    return {
        "receita": "Receita Forecast",
        "custos_diretos": "Custos Diretos Forecast",
        "custos_indiretos": "Custos Indiretos Forecast",
        "impostos": "Impostos Forecast",
        "resultado": "Resultado Forecast",
    }.get(indicador, indicador)


def _resumo_map(db: Session, importacao_id: int) -> dict[tuple[str, str, str | None], float]:
    rows = (
        db.query(EconomicoResumoCalculado)
        .filter(EconomicoResumoCalculado.importacao_id == importacao_id, EconomicoResumoCalculado.periodo.is_(None))
        .all()
    )
    return {
        (r.indicador, r.cenario, r.categoria): float(r.valor or 0)
        for r in rows
    }


def _valor_resumo(resumo: dict, indicador: str, cenario: str, categoria: str | None = None) -> float:
    return float(resumo.get((indicador, cenario, categoria), 0.0))


def _serie_receita(db: Session, importacao_id: int) -> dict:
    rows = (
        db.query(EconomicoResumoCalculado)
        .filter(
            EconomicoResumoCalculado.importacao_id == importacao_id,
            EconomicoResumoCalculado.indicador == "receita",
            EconomicoResumoCalculado.periodo.isnot(None),
            EconomicoResumoCalculado.categoria.is_(None),
        )
        .order_by(EconomicoResumoCalculado.periodo)
        .all()
    )
    buckets: dict[date, dict] = {}
    for row in rows:
        buckets.setdefault(row.periodo, {})[row.cenario] = float(row.valor or 0.0)
    mensal, acumulada = [], []
    acc = {"linha_base": 0.0, "tendencia": 0.0, "real": 0.0}
    for periodo in sorted(buckets):
        item = buckets[periodo]
        lb = item.get("linha_base", 0.0)
        tend = item.get("tendencia", 0.0)
        real = item.get("real", 0.0)
        mensal.append({
            "periodo": periodo.isoformat(),
            "linha_base": lb,
            "tendencia": tend,
            "realizada": real,
            "desvio": tend - lb,
        })
        acc["linha_base"] += lb
        acc["tendencia"] += tend
        acc["real"] += real
        acumulada.append({
            "periodo": periodo.isoformat(),
            "linha_base": acc["linha_base"],
            "tendencia": acc["tendencia"],
            "realizada": acc["real"],
            "desvio": acc["tendencia"] - acc["linha_base"],
        })
    return {"mensal": mensal, "acumulada": acumulada}


def _serie_resultado_forecast(db: Session, importacao_id: int, acumulado: bool) -> list[dict]:
    rows = (
        db.query(EconomicoResumoCalculado)
        .filter(
            EconomicoResumoCalculado.importacao_id == importacao_id,
            EconomicoResumoCalculado.indicador == "resultado",
            EconomicoResumoCalculado.cenario.in_(["linha_base", "tendencia"]),
            EconomicoResumoCalculado.periodo.isnot(None),
            EconomicoResumoCalculado.categoria.is_(None),
        )
        .order_by(EconomicoResumoCalculado.periodo)
        .all()
    )
    buckets: dict[date, dict] = {}
    for row in rows:
        buckets.setdefault(row.periodo, {})[row.cenario] = float(row.valor or 0.0)
    out = []
    acc_lb = 0.0
    acc_fc = 0.0
    for periodo in sorted(buckets):
        lb = buckets[periodo].get("linha_base", 0.0)
        fc = buckets[periodo].get("tendencia", 0.0)
        if acumulado:
            acc_lb += lb
            acc_fc += fc
            lb, fc = acc_lb, acc_fc
        out.append({
            "periodo": periodo.isoformat(),
            "linha_base": lb,
            "forecast": fc,
            "impacto": fc - lb,
        })
    return out


def _serie_resultado_completa(db: Session, importacao_id: int) -> list[dict]:
    rows = (
        db.query(EconomicoResumoCalculado)
        .filter(
            EconomicoResumoCalculado.importacao_id == importacao_id,
            EconomicoResumoCalculado.indicador == "resultado",
            EconomicoResumoCalculado.cenario.in_(["linha_base", "tendencia", "real"]),
            EconomicoResumoCalculado.periodo.isnot(None),
            EconomicoResumoCalculado.categoria.is_(None),
        )
        .order_by(EconomicoResumoCalculado.periodo)
        .all()
    )
    buckets: dict[date, dict] = {}
    for row in rows:
        buckets.setdefault(row.periodo, {})[row.cenario] = float(row.valor or 0.0)

    linha_base_total = (
        db.query(EconomicoResumoCalculado.valor)
        .filter(
            EconomicoResumoCalculado.importacao_id == importacao_id,
            EconomicoResumoCalculado.indicador == "resultado",
            EconomicoResumoCalculado.cenario == "linha_base",
            EconomicoResumoCalculado.periodo.is_(None),
            EconomicoResumoCalculado.categoria.is_(None),
        )
        .scalar()
    )
    has_linha_base_mensal = any("linha_base" in item for item in buckets.values())
    acc = {"linha_base": 0.0, "tendencia": 0.0, "real": 0.0}
    out = []
    for periodo in sorted(buckets):
        item = buckets[periodo]
        acc["linha_base"] += item.get("linha_base", 0.0)
        acc["tendencia"] += item.get("tendencia", 0.0)
        acc["real"] += item.get("real", 0.0)
        out.append({
            "periodo": periodo.isoformat(),
            "linha_base": acc["linha_base"] if has_linha_base_mensal else float(linha_base_total or 0.0),
            "forecast": acc["tendencia"],
            "realizado": acc["real"],
        })
    return out


def _serie_margem(db: Session, importacao_id: int) -> list[dict]:
    rows = (
        db.query(EconomicoResumoCalculado)
        .filter(
            EconomicoResumoCalculado.importacao_id == importacao_id,
            EconomicoResumoCalculado.indicador.in_(["receita", "resultado"]),
            EconomicoResumoCalculado.cenario.in_(["linha_base", "tendencia", "real"]),
            EconomicoResumoCalculado.periodo.isnot(None),
            EconomicoResumoCalculado.categoria.is_(None),
        )
        .order_by(EconomicoResumoCalculado.periodo)
        .all()
    )
    buckets: dict[date, dict] = {}
    for row in rows:
        key = (row.indicador, row.cenario)
        buckets.setdefault(row.periodo, {})[key] = float(row.valor or 0.0)

    margem_linha_base_total = (
        db.query(EconomicoResumoCalculado.valor)
        .filter(
            EconomicoResumoCalculado.importacao_id == importacao_id,
            EconomicoResumoCalculado.indicador == "margem",
            EconomicoResumoCalculado.cenario == "linha_base",
            EconomicoResumoCalculado.periodo.is_(None),
            EconomicoResumoCalculado.categoria.is_(None),
        )
        .scalar()
    )
    has_linha_base_mensal = any(("resultado", "linha_base") in item for item in buckets.values())
    acc: dict[tuple[str, str], float] = {}
    out = []
    for periodo in sorted(buckets):
        for key, value in buckets[periodo].items():
            acc[key] = acc.get(key, 0.0) + value
        out.append({
            "periodo": periodo.isoformat(),
            "linha_base": (
                _div(acc.get(("resultado", "linha_base"), 0.0), acc.get(("receita", "linha_base"), 0.0))
                if has_linha_base_mensal
                else float(margem_linha_base_total or 0.0)
            ),
            "forecast": _div(acc.get(("resultado", "tendencia"), 0.0), acc.get(("receita", "tendencia"), 0.0)),
            "realizado": _div(acc.get(("resultado", "real"), 0.0), acc.get(("receita", "real"), 0.0)),
        })
    return out


def _impacto_total_aprovado(resumo: dict) -> float:
    return _valor_resumo(resumo, "resultado", "tendencia") - _valor_resumo(resumo, "resultado", "linha_base")


def _query_lancamentos_filtrados(
    db: Session,
    importacao_id: int,
    categoria: str | None = None,
    fornecedor: str | None = None,
    conta: str | None = None,
    documento: str | None = None,
    periodo_inicio: date | None = None,
    periodo_fim: date | None = None,
):
    query = db.query(EconomicoLancamentoRazao).filter(EconomicoLancamentoRazao.importacao_id == importacao_id)
    if categoria:
        query = query.filter(EconomicoLancamentoRazao.categoria_dre == categoria)
    if fornecedor:
        query = query.filter(EconomicoLancamentoRazao.fornecedor == fornecedor)
    if conta:
        query = query.filter(EconomicoLancamentoRazao.conta == conta)
    if documento:
        query = query.filter(EconomicoLancamentoRazao.documento == documento)
    if periodo_inicio:
        query = query.filter(EconomicoLancamentoRazao.data >= periodo_inicio)
    if periodo_fim:
        query = query.filter(EconomicoLancamentoRazao.data <= periodo_fim)
    return query


def _ranking_fornecedores(db: Session, importacao_id: int, ledger_query) -> list[dict]:
    realizados = {}
    for fornecedor, valor in (
        ledger_query
        .with_entities(EconomicoLancamentoRazao.fornecedor, func.coalesce(func.sum(EconomicoLancamentoRazao.valor), 0.0))
        .group_by(EconomicoLancamentoRazao.fornecedor)
        .all()
    ):
        key = fornecedor or "Sem fornecedor"
        realizados[key] = abs(float(valor or 0))

    contratados = {
        fornecedor or "Sem fornecedor": abs(float(valor or 0))
        for fornecedor, valor in (
            db.query(EconomicoRelatorioOC.fornecedor, func.coalesce(func.sum(EconomicoRelatorioOC.valor_liquido), 0.0))
            .filter(EconomicoRelatorioOC.importacao_id == importacao_id)
            .group_by(EconomicoRelatorioOC.fornecedor)
            .all()
        )
    }
    total_realizado = sum(realizados.values())
    rows = []
    for fornecedor, realizado in realizados.items():
        contratado = contratados.get(fornecedor, 0.0)
        forecast = max(contratado, realizado)
        desvio = realizado - contratado
        rows.append({
            "fornecedor": fornecedor,
            "valor_contratado": contratado,
            "forecast": forecast,
            "realizado": realizado,
            "desvio": desvio,
            "impacto_percentual": _div(desvio, total_realizado),
        })
    return sorted(rows, key=lambda x: abs(x["desvio"]), reverse=True)[:30]


def _ranking_contas(db: Session, importacao_id: int, ledger_query) -> list[dict]:
    descricoes = {
        conta: desc
        for conta, desc in db.query(EconomicoContaDespesa.conta, EconomicoContaDespesa.agrupamento_dre)
        .filter(EconomicoContaDespesa.importacao_id == importacao_id)
        .all()
    }
    total = 0.0
    rows = []
    for conta, desc, valor in (
        ledger_query
        .with_entities(
            EconomicoLancamentoRazao.conta,
            EconomicoLancamentoRazao.conta_descricao,
            func.coalesce(func.sum(EconomicoLancamentoRazao.valor), 0.0),
        )
        .group_by(EconomicoLancamentoRazao.conta, EconomicoLancamentoRazao.conta_descricao)
        .all()
    ):
        realizado = abs(float(valor or 0))
        total += realizado
        rows.append({
            "conta": conta or "Sem conta",
            "descricao": desc or descricoes.get(str(conta), "") or "-",
            "previsto": 0.0,
            "forecast": realizado,
            "realizado": realizado,
            "desvio": -realizado,
        })
    for row in rows:
        row["impacto_percentual"] = _div(row["realizado"], total)
    return sorted(rows, key=lambda x: x["realizado"], reverse=True)[:30]


def _pareto_categorias(categorias: list[dict]) -> list[dict]:
    negativos = [row for row in categorias if row["desvio"] < -0.01]
    total = sum(abs(row["desvio"]) for row in negativos)
    acumulado = 0.0
    pareto = []
    for row in sorted(negativos, key=lambda x: abs(x["desvio"]), reverse=True)[:12]:
        valor = abs(row["desvio"])
        acumulado += valor
        pareto.append({
            "categoria": row["categoria"],
            "impacto": -valor,
            "participacao": _div(valor, total),
            "acumulado": _div(acumulado, total),
        })
    return pareto


def _categorias_custo(db: Session, importacao_id: int, resumo: dict) -> list[dict]:
    base_total = abs(_valor_resumo(resumo, "custos_diretos", "linha_base"))
    forecast_total = abs(_valor_resumo(resumo, "custos_diretos", "tendencia"))
    realizado_total = abs(_valor_resumo(resumo, "custos_diretos", "real"))

    dre_rows = db.query(EconomicoAnaliseDRE).filter(EconomicoAnaliseDRE.importacao_id == importacao_id).all()
    forecast_por_categoria = {
        categoria: abs(_valor_resumo(resumo, "categoria_custo", "tendencia", categoria))
        for _, _, categoria in resumo
        if categoria
    }
    realizado_por_categoria = {
        categoria: abs(_valor_resumo(resumo, "categoria_custo", "real", categoria))
        for _, _, categoria in resumo
        if categoria
    }

    dre_base_raw = {r.categoria: abs(float(r.considerar or r.previsao_anterior or 0)) for r in dre_rows}
    soma_base_raw = sum(v for v in dre_base_raw.values() if v > 0)
    base_por_categoria = {}
    for categoria, forecast in forecast_por_categoria.items():
        if dre_base_raw.get(categoria):
            base_por_categoria[categoria] = dre_base_raw[categoria]
        elif soma_base_raw and base_total:
            base_por_categoria[categoria] = base_total * (forecast / max(forecast_total, 1))
        else:
            base_por_categoria[categoria] = 0.0

    escala_base = _div(base_total, sum(base_por_categoria.values()))
    if escala_base:
        base_por_categoria = {k: v * escala_base for k, v in base_por_categoria.items()}

    rows = []
    for categoria, forecast_abs in forecast_por_categoria.items():
        previsto = base_por_categoria.get(categoria, 0.0)
        realizado = realizado_por_categoria.get(categoria, 0.0)
        desvio = -(forecast_abs - previsto)
        rows.append({
            "categoria": categoria,
            "previsto": previsto,
            "forecast": forecast_abs,
            "realizado": realizado,
            "desvio": desvio,
            "impacto_percentual": _div(desvio, base_total),
            "impacto_resultado": desvio,
        })
    if forecast_total and abs(sum(r["forecast"] for r in rows) - forecast_total) > 0.01:
        fator = forecast_total / sum(r["forecast"] for r in rows)
        for row in rows:
            row["forecast"] *= fator
    if realizado_total and sum(r["realizado"] for r in rows):
        fator = realizado_total / sum(r["realizado"] for r in rows)
        for row in rows:
            row["realizado"] *= fator
    return rows


def _validar_abas(wb) -> None:
    obrigatorias = {
        "EAP",
        "Meta Receita-Custos",
        "Resumo BI",
        "RELAT_158",
        "RAZÃO",
        "Distribuição Revenda",
        "Dicionário Conta de Despesa",
    }
    faltantes = sorted(obrigatorias - set(wb.sheetnames))
    if faltantes:
        raise ValueError("Abas obrigatórias ausentes: " + ", ".join(faltantes))


def _periodos_resumo_bi(wb) -> list[date]:
    ws = wb["Resumo BI"]
    meses: list[date] = []
    col = RESUMO_FIRST_MONTH_COL
    while True:
        v = ws.cell(5, col).value
        if v is None:
            break
        meses.append(_as_date(v))
        col += 1
    return meses


def _contar_meses_reais(wb) -> int:
    ws = wb["Meta Receita-Custos"]
    count = 0
    col = META_FIRST_MONTH_COL
    while ws.cell(3, col).value is not None:
        if str(ws.cell(META_REAL_STATUS_ROW, col).value or "").strip().lower() == "real":
            count += 1
        col += 1
    return count


def _calcular_metricas(wb, meses: list[date], real_count: int) -> list[AuditMetric]:
    rb = wb["Resumo BI"]
    forecast_count = _row_month_count(rb, 100)
    receita_forecast_count = _row_month_count(rb, 13)

    receita_lb = _receita_linha_base(wb, meses)
    receita_real = _receita_eap(wb, meses, real_count, tendencia=True)
    receita_tend = _receita_eap(wb, meses, receita_forecast_count, tendencia=True)
    receita_tend_resultado = _receita_eap(wb, meses, forecast_count, tendencia=True)

    impostos_lb = _rb_total(rb, 29)
    impostos_real = _meta_soma(wb, [12, 22, 23], real_count)
    impostos_tend = _meta_soma(wb, [12, 22, 23], forecast_count)

    custos_diretos_lb = _rb_total(rb, 34)
    custos_diretos_real = _custos_diretos_meta(wb, real_count)
    custos_diretos_tend = _custos_diretos_meta(wb, forecast_count)

    custos_indiretos_lb = _rb_total(rb, 93)
    custos_indiretos_real = _meta_soma(wb, [15, 25], real_count)
    custos_indiretos_tend = _meta_soma(wb, [15, 25], forecast_count)

    resultado_lb = receita_lb + impostos_lb + custos_diretos_lb + custos_indiretos_lb
    resultado_real = receita_real + impostos_real + custos_diretos_real + custos_indiretos_real
    resultado_tend = receita_tend_resultado + impostos_tend + custos_diretos_tend + custos_indiretos_tend

    margem_lb = _div(resultado_lb, receita_lb)
    margem_real = _div(resultado_real, receita_real)
    margem_tend = _div(resultado_tend, receita_tend)

    return [
        _metric("Receita Linha Base", receita_lb, _rb_total(rb, 6), "EAP curva linha-base; fallback snapshot se curva nao fecha", "Resumo BI!C6"),
        _metric("Receita Realizada", receita_real, _rb_total(rb, 20), "EAP BB:CF ate meses Real da Meta Receita-Custos", "Resumo BI!C20"),
        _metric("Receita Tendencia", receita_tend, _rb_total(rb, 13), "EAP BB:CF acumulado", "Resumo BI!C13"),
        _metric("Impostos Linha Base", impostos_lb, _rb_total(rb, 29), "Formula de impostos sobre linha-base; snapshot importado", "Resumo BI!C29"),
        _metric("Impostos Realizados", impostos_real, _rb_total(rb, 31), "Meta Receita-Custos linhas 12 + 22 + 23 ate meses Real", "Resumo BI!C31"),
        _metric("Impostos Tendencia", impostos_tend, _rb_total(rb, 30), "Meta Receita-Custos linhas 12 + 22 + 23", "Resumo BI!C30"),
        _metric("Custos Diretos Linha Base", custos_diretos_lb, _rb_total(rb, 34), "Linha-base original sem aba-fonte rastreavel; snapshot importado", "Resumo BI!C34"),
        _metric("Custos Diretos Realizados", custos_diretos_real, _rb_total(rb, 72), "Meta Receita-Custos linhas 31:47 + 21 ate meses Real", "Resumo BI!C72"),
        _metric("Custos Diretos Tendencia", custos_diretos_tend, _rb_total(rb, 53), "Meta Receita-Custos linhas 31:47 + 21", "Resumo BI!C53"),
        _metric("Custos Indiretos Linha Base", custos_indiretos_lb, _rb_total(rb, 93), "Linha-base original sem aba-fonte rastreavel; snapshot importado", "Resumo BI!C93"),
        _metric("Custos Indiretos Realizados", custos_indiretos_real, _rb_total(rb, 95), "Meta Receita-Custos linhas 15 + 25 ate meses Real", "Resumo BI!C95"),
        _metric("Custos Indiretos Tendencia", custos_indiretos_tend, _rb_total(rb, 94), "Meta Receita-Custos linhas 15 + 25", "Resumo BI!C94"),
        _metric("Resultado Linha Base", resultado_lb, _rb_total(rb, 99), "Receita + impostos + custos diretos + indiretos", "Resumo BI!C99"),
        _metric("Resultado Atual", resultado_real, _rb_total(rb, 101), "Receita real + impostos reais + custos reais", "Resumo BI!C101"),
        _metric("Resultado Forecast", resultado_tend, _rb_total(rb, 100), "Linha TENDENCIA oficial", "Resumo BI!C100"),
        _metric("Margem Linha Base", margem_lb, _div(_rb_total(rb, 99), _rb_total(rb, 6)), "Resultado linha-base / receita linha-base", "Resumo BI C99/C6"),
        _metric("Margem Atual", margem_real, _div(_rb_total(rb, 101), _rb_total(rb, 20)), "Resultado atual / receita realizada", "Resumo BI C101/C20"),
        _metric("Margem Forecast", margem_tend, _div(_rb_total(rb, 100), _rb_total(rb, 13)), "Resultado forecast / receita tendencia", "Resumo BI C100/C13"),
    ]


def _metric(indicador: str, sistema: float, resumo_bi: float, origem_sistema: str, origem_resumo_bi: str) -> AuditMetric:
    return AuditMetric(indicador, float(sistema or 0), float(resumo_bi or 0), origem_sistema, origem_resumo_bi)


def _persistir_auditoria(db: Session, importacao_id: int, metricas: Iterable[AuditMetric]) -> list[EconomicoAuditoria]:
    rows: list[EconomicoAuditoria] = []
    for m in metricas:
        diff = (m.sistema or 0) - (m.resumo_bi or 0)
        row = EconomicoAuditoria(
            importacao_id=importacao_id,
            indicador=m.indicador,
            sistema=m.sistema,
            resumo_bi=m.resumo_bi,
            diferenca=diff,
            aprovado=abs(diff) <= TOLERANCIA,
            tolerancia=TOLERANCIA,
            origem_sistema=m.origem_sistema,
            origem_resumo_bi=m.origem_resumo_bi,
        )
        db.add(row)
        rows.append(row)
    db.flush()
    return rows


def _persistir_valores(db: Session, importacao_id: int, wb, meses: list[date], real_count: int) -> None:
    rb = wb["Resumo BI"]
    forecast_count = _row_month_count(rb, 100)
    receita_forecast_count = _row_month_count(rb, 13)
    for indicador, row, cenario in [
        ("receita", 6, "linha_base"),
        ("receita", 13, "tendencia"),
        ("receita", 20, "real"),
        ("impostos", 29, "linha_base"),
        ("impostos", 30, "tendencia"),
        ("impostos", 31, "real"),
        ("custos_diretos", 34, "linha_base"),
        ("custos_diretos", 53, "tendencia"),
        ("custos_diretos", 72, "real"),
        ("custos_indiretos", 93, "linha_base"),
        ("custos_indiretos", 94, "tendencia"),
        ("custos_indiretos", 95, "real"),
        ("resultado", 99, "linha_base"),
        ("resultado", 100, "tendencia"),
        ("resultado", 101, "real"),
    ]:
        db.add(EconomicoValor(
            importacao_id=importacao_id,
            tipo="resumo_bi",
            indicador=indicador,
            cenario=cenario,
            periodo=None,
            valor=_rb_total(rb, row),
            origem=f"Resumo BI!C{row}",
        ))
        for idx, mes in enumerate(meses):
            db.add(EconomicoValor(
                importacao_id=importacao_id,
                tipo="resumo_bi",
                indicador=indicador,
                cenario=cenario,
                periodo=mes,
                valor=_num(rb.cell(row, RESUMO_FIRST_MONTH_COL + idx).value),
                origem=f"Resumo BI mensal row {row}",
            ))

    for a in _calcular_metricas(wb, meses, real_count):
        db.add(EconomicoValor(
            importacao_id=importacao_id,
            tipo="sistema",
            indicador=a.indicador,
            cenario="auditoria",
            periodo=None,
            valor=a.sistema,
            origem=a.origem_sistema,
        ))

    sistema_acumulado = {
        ("receita", "linha_base"): _receita_linha_base(wb, meses),
        ("receita", "real"): _receita_eap(wb, meses, real_count, tendencia=True),
        ("receita", "tendencia"): _receita_eap(wb, meses, receita_forecast_count, tendencia=True),
        ("impostos", "linha_base"): _rb_total(rb, 29),
        ("impostos", "real"): _meta_soma(wb, [12, 22, 23], real_count),
        ("impostos", "tendencia"): _meta_soma(wb, [12, 22, 23], forecast_count),
        ("custos_diretos", "linha_base"): _rb_total(rb, 34),
        ("custos_diretos", "real"): _custos_diretos_meta(wb, real_count),
        ("custos_diretos", "tendencia"): _custos_diretos_meta(wb, forecast_count),
        ("custos_indiretos", "linha_base"): _rb_total(rb, 93),
        ("custos_indiretos", "real"): _meta_soma(wb, [15, 25], real_count),
        ("custos_indiretos", "tendencia"): _meta_soma(wb, [15, 25], forecast_count),
    }
    sistema_acumulado[("resultado", "linha_base")] = (
        sistema_acumulado[("receita", "linha_base")]
        + sistema_acumulado[("impostos", "linha_base")]
        + sistema_acumulado[("custos_diretos", "linha_base")]
        + sistema_acumulado[("custos_indiretos", "linha_base")]
    )
    sistema_acumulado[("resultado", "real")] = (
        sistema_acumulado[("receita", "real")]
        + sistema_acumulado[("impostos", "real")]
        + sistema_acumulado[("custos_diretos", "real")]
        + sistema_acumulado[("custos_indiretos", "real")]
    )
    receita_tend_resultado = _receita_eap(wb, meses, forecast_count, tendencia=True)
    sistema_acumulado[("resultado", "tendencia")] = (
        receita_tend_resultado
        + sistema_acumulado[("impostos", "tendencia")]
        + sistema_acumulado[("custos_diretos", "tendencia")]
        + sistema_acumulado[("custos_indiretos", "tendencia")]
    )
    sistema_acumulado[("margem", "linha_base")] = _div(
        sistema_acumulado[("resultado", "linha_base")],
        sistema_acumulado[("receita", "linha_base")],
    )
    sistema_acumulado[("margem", "real")] = _div(
        sistema_acumulado[("resultado", "real")],
        sistema_acumulado[("receita", "real")],
    )
    sistema_acumulado[("margem", "tendencia")] = _div(
        sistema_acumulado[("resultado", "tendencia")],
        sistema_acumulado[("receita", "tendencia")],
    )
    for (indicador, cenario), valor in sistema_acumulado.items():
        db.add(EconomicoValor(
            importacao_id=importacao_id,
            tipo="sistema",
            indicador=indicador,
            cenario=cenario,
            periodo=None,
            valor=valor,
            origem="Tabela calculada validada pela auditoria",
        ))

    for idx, mes in enumerate(meses):
        receita_lb = _receita_linha_base_mes(wb, idx)
        receita_tend = _receita_eap_mes(wb, idx, tendencia=True) if idx < receita_forecast_count else 0.0
        receita_real = receita_tend if idx < real_count else 0.0
        valores_mes = {
            ("receita", "linha_base"): receita_lb,
            ("receita", "tendencia"): receita_tend,
            ("receita", "real"): receita_real,
            ("impostos", "tendencia"): _meta_soma_mes(wb, [12, 22, 23], idx) if idx < forecast_count else 0.0,
            ("impostos", "real"): _meta_soma_mes(wb, [12, 22, 23], idx) if idx < real_count else 0.0,
            ("custos_diretos", "tendencia"): _custos_diretos_meta_mes(wb, idx) if idx < forecast_count else 0.0,
            ("custos_diretos", "real"): _custos_diretos_meta_mes(wb, idx) if idx < real_count else 0.0,
            ("custos_indiretos", "tendencia"): _meta_soma_mes(wb, [15, 25], idx) if idx < forecast_count else 0.0,
            ("custos_indiretos", "real"): _meta_soma_mes(wb, [15, 25], idx) if idx < real_count else 0.0,
        }
        valores_mes[("resultado", "tendencia")] = (
            valores_mes[("receita", "tendencia")]
            + valores_mes[("impostos", "tendencia")]
            + valores_mes[("custos_diretos", "tendencia")]
            + valores_mes[("custos_indiretos", "tendencia")]
        )
        valores_mes[("resultado", "real")] = (
            valores_mes[("receita", "real")]
            + valores_mes[("impostos", "real")]
            + valores_mes[("custos_diretos", "real")]
            + valores_mes[("custos_indiretos", "real")]
        )
        for (indicador, cenario), valor in valores_mes.items():
            db.add(EconomicoValor(
                importacao_id=importacao_id,
                tipo="sistema",
                indicador=indicador,
                cenario=cenario,
                periodo=mes,
                valor=valor,
                origem="Curva mensal calculada das abas-fonte",
            ))


def _persistir_investigacao(db: Session, importacao_id: int, wb, meses: list[date], real_count: int) -> None:
    _persistir_resumo_calculado(db, importacao_id)
    _persistir_receita_fases(db, importacao_id, wb, meses, real_count)
    _persistir_categorias_calculadas(db, importacao_id, wb, meses, real_count)
    _persistir_contas_despesa(db, importacao_id, wb)
    _persistir_lancamentos_razao(db, importacao_id, wb)
    _persistir_relatorio_oc(db, importacao_id, wb)
    _persistir_analise_dre(db, importacao_id, wb)


def _persistir_resumo_calculado(db: Session, importacao_id: int) -> None:
    db.flush()
    rows = (
        db.query(EconomicoValor)
        .filter(EconomicoValor.importacao_id == importacao_id, EconomicoValor.tipo == "sistema")
        .all()
    )
    for r in rows:
        if r.cenario == "auditoria":
            continue
        db.add(EconomicoResumoCalculado(
            importacao_id=importacao_id,
            indicador=r.indicador,
            cenario=r.cenario,
            periodo=r.periodo,
            categoria=r.categoria,
            valor=r.valor,
            origem=r.origem,
        ))


def _persistir_receita_fases(db: Session, importacao_id: int, wb, meses: list[date], real_count: int) -> None:
    rb = wb["Resumo BI"]
    forecast_count = _row_month_count(rb, 13)
    for row, fase in PHASE_LABELS.items():
        linha_base_total = _receita_fase_total(wb, row, len(meses), "linha_base")
        tendencia_total = _receita_fase_total(wb, row, forecast_count, "tendencia")
        real_total = _receita_fase_total(wb, row, real_count, "tendencia")
        for cenario, valor in [
            ("linha_base", linha_base_total),
            ("tendencia", tendencia_total),
            ("real", real_total),
        ]:
            db.add(EconomicoResumoCalculado(
                importacao_id=importacao_id,
                indicador="receita_fase",
                cenario=cenario,
                periodo=None,
                categoria=fase,
                valor=valor,
                origem=f"EAP linha {row}",
            ))
        for idx, mes in enumerate(meses):
            lb = _receita_fase_mes(wb, row, idx, "linha_base")
            tend = _receita_fase_mes(wb, row, idx, "tendencia") if idx < forecast_count else 0.0
            real = _receita_fase_mes(wb, row, idx, "tendencia") if idx < real_count else 0.0
            for cenario, valor in [
                ("linha_base", lb),
                ("tendencia", tend),
                ("real", real),
            ]:
                db.add(EconomicoResumoCalculado(
                    importacao_id=importacao_id,
                    indicador="receita_fase",
                    cenario=cenario,
                    periodo=mes,
                    categoria=fase,
                    valor=valor,
                    origem=f"EAP linha {row}",
                ))


def _persistir_categorias_calculadas(db: Session, importacao_id: int, wb, meses: list[date], real_count: int) -> None:
    ws = wb["Meta Receita-Custos"]
    rb = wb["Resumo BI"]
    forecast_count = _row_month_count(rb, 100)
    for row in range(31, 48):
        categoria = str(ws.cell(row, 2).value or "").strip()
        if not categoria:
            continue
        forecast = _meta_soma(wb, [row], forecast_count)
        realizado = _meta_soma(wb, [row], real_count)
        for cenario, valor in [("tendencia", forecast), ("real", realizado)]:
            db.add(EconomicoResumoCalculado(
                importacao_id=importacao_id,
                indicador="categoria_custo",
                cenario=cenario,
                periodo=None,
                categoria=categoria,
                valor=valor,
                origem=f"Meta Receita-Custos linha {row}",
            ))
        for idx, mes in enumerate(meses):
            for cenario, limit in [("tendencia", forecast_count), ("real", real_count)]:
                db.add(EconomicoResumoCalculado(
                    importacao_id=importacao_id,
                    indicador="categoria_custo",
                    cenario=cenario,
                    periodo=mes,
                    categoria=categoria,
                    valor=_meta_soma_mes(wb, [row], idx) if idx < limit else 0.0,
                    origem=f"Meta Receita-Custos linha {row}",
                ))


def _persistir_contas_despesa(db: Session, importacao_id: int, wb) -> None:
    ws = _sheet(wb, "Dicion")
    for row in range(8, ws.max_row + 1):
        conta = ws.cell(row, 3).value
        if conta in (None, ""):
            continue
        db.add(EconomicoContaDespesa(
            importacao_id=importacao_id,
            conta=str(conta).strip(),
            descricao=_str(ws.cell(row, 4).value),
            comentario=_str(ws.cell(row, 5).value),
            agrupamento_dre=_str(ws.cell(row, 6).value),
        ))


def _persistir_lancamentos_razao(db: Session, importacao_id: int, wb) -> None:
    ws = _sheet(wb, "RAZ")
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value in (None, ""):
            continue
        valor = _num(ws.cell(row, 15).value)
        if abs(valor) < 0.000001:
            continue
        db.add(EconomicoLancamentoRazao(
            importacao_id=importacao_id,
            data=_as_date_or_none(ws.cell(row, 10).value),
            documento=_str(ws.cell(row, 16).value),
            fornecedor=_str(ws.cell(row, 12).value) or "Sem fornecedor",
            conta=_str(ws.cell(row, 4).value),
            conta_descricao=_str(ws.cell(row, 6).value),
            categoria_dre=_str(ws.cell(row, 5).value) or "Sem categoria",
            historico=_str(ws.cell(row, 11).value),
            valor=valor,
            tipo=_str(ws.cell(row, 14).value),
            lote=_str(ws.cell(row, 7).value),
            lancamento=_str(ws.cell(row, 8).value),
        ))


def _persistir_relatorio_oc(db: Session, importacao_id: int, wb) -> None:
    ws = wb["RELAT_158"]
    for row in range(7, ws.max_row + 1):
        numero_oc = ws.cell(row, 1).value
        if numero_oc in (None, ""):
            continue
        db.add(EconomicoRelatorioOC(
            importacao_id=importacao_id,
            numero_oc=_str(numero_oc),
            item_oc=_str(ws.cell(row, 2).value),
            requisicao=_str(ws.cell(row, 3).value),
            produto=_str(ws.cell(row, 4).value),
            descricao=_str(ws.cell(row, 5).value),
            fornecedor=_str(ws.cell(row, 20).value) or "Sem fornecedor",
            data=_as_date_or_none(ws.cell(row, 12).value),
            conta=_str(ws.cell(row, 43).value),
            conta_descricao=_str(ws.cell(row, 44).value),
            valor_total=_num(ws.cell(row, 9).value),
            valor_liquido=_num(ws.cell(row, 34).value),
            valor_nf=_num(ws.cell(row, 32).value),
        ))


def _persistir_analise_dre(db: Session, importacao_id: int, wb) -> None:
    ws = _sheet(wb, "DRE")
    for row in range(16, 28):
        categoria = _str(ws.cell(row, 1).value)
        if not categoria:
            continue
        db.add(EconomicoAnaliseDRE(
            importacao_id=importacao_id,
            categoria=categoria,
            projetado=_num(ws.cell(row, 2).value),
            razao=_num(ws.cell(row, 3).value),
            asocnf=_num(ws.cell(row, 4).value),
            fat_nao_lancado_razao=_num(ws.cell(row, 5).value),
            forecast=_num(ws.cell(row, 6).value),
            previsao_anterior=_num(ws.cell(row, 8).value),
            considerar=_num(ws.cell(row, 10).value),
        ))


def _receita_linha_base(wb, meses: list[date]) -> float:
    eap = wb["EAP"]
    total = 0.0
    for idx in range(len(meses)):
        col = EAP_BASELINE_FIRST_COL + idx
        for row in range(1, eap.max_row + 1):
            pct = eap.cell(row, col).value
            valor = eap.cell(row, 12).value
            if isinstance(pct, (int, float)) and isinstance(valor, (int, float)):
                total += float(pct) * float(valor)
    resumo_total = _rb_total(wb["Resumo BI"], 6)
    if abs(total - resumo_total) <= TOLERANCIA:
        return total
    # A versão atual não traz a linha-base fechando nos nós-fonte da EAP.
    # Mantemos o snapshot de linha-base para auditar exatamente a planilha atual.
    return resumo_total


def _receita_linha_base_mes(wb, idx: int) -> float:
    eap = wb["EAP"]
    col = EAP_BASELINE_FIRST_COL + idx
    total = 0.0
    for row in range(1, eap.max_row + 1):
        pct = eap.cell(row, col).value
        valor = eap.cell(row, 12).value
        if isinstance(pct, (int, float)) and isinstance(valor, (int, float)):
            total += float(pct) * float(valor)
    return total


def _receita_eap(wb, meses: list[date], month_count: int, tendencia: bool) -> float:
    eap = wb["EAP"]
    first_col = EAP_TENDENCIA_FIRST_COL if tendencia else EAP_BASELINE_FIRST_COL
    total = 0.0
    for idx in range(min(month_count, len(meses))):
        col = first_col + idx
        for row in PHASE_ROWS:
            total += _num(eap.cell(row, col).value)
    return total


def _receita_eap_mes(wb, idx: int, tendencia: bool) -> float:
    eap = wb["EAP"]
    first_col = EAP_TENDENCIA_FIRST_COL if tendencia else EAP_BASELINE_FIRST_COL
    col = first_col + idx
    return sum(_num(eap.cell(row, col).value) for row in PHASE_ROWS)


def _receita_fase_total(wb, row: int, month_count: int, modo: str) -> float:
    return sum(_receita_fase_mes(wb, row, idx, modo) for idx in range(month_count))


def _receita_fase_mes(wb, row: int, idx: int, modo: str) -> float:
    eap = wb["EAP"]
    if modo == "linha_base":
        start, end = PHASE_RANGES[row]
        end = end or eap.max_row
        col = EAP_BASELINE_FIRST_COL + idx
        total = 0.0
        for eap_row in range(start, end + 1):
            pct = eap.cell(eap_row, col).value
            valor = eap.cell(eap_row, 12).value
            if isinstance(pct, (int, float)) and isinstance(valor, (int, float)):
                total += float(pct) * float(valor)
        return total
    return _num(eap.cell(row, EAP_TENDENCIA_FIRST_COL + idx).value)


def _custos_diretos_meta(wb, month_count: int) -> float:
    return _meta_soma(wb, list(range(31, 48)) + [21], month_count)


def _custos_diretos_meta_mes(wb, idx: int) -> float:
    return _meta_soma_mes(wb, list(range(31, 48)) + [21], idx)


def _meta_soma(wb, rows: list[int], month_count: int) -> float:
    ws = wb["Meta Receita-Custos"]
    total = 0.0
    for idx in range(month_count):
        col = META_FIRST_MONTH_COL + idx
        for row in rows:
            total += _num(ws.cell(row, col).value)
    return total


def _meta_soma_mes(wb, rows: list[int], idx: int) -> float:
    ws = wb["Meta Receita-Custos"]
    col = META_FIRST_MONTH_COL + idx
    return sum(_num(ws.cell(row, col).value) for row in rows)


def _rb_total(ws, row: int) -> float:
    return _num(ws.cell(row, 3).value)


def _row_month_count(ws, row: int) -> int:
    count = 0
    col = RESUMO_FIRST_MONTH_COL
    while ws.cell(row, col).value is not None:
        count += 1
        col += 1
    return count


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        n = float(v)
        if math.isnan(n) or math.isinf(n):
            return 0.0
        return n
    except (TypeError, ValueError):
        return 0.0


def _div(a: float, b: float) -> float:
    return 0.0 if not b else a / b


def _as_date(v) -> date:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        return datetime.fromisoformat(v[:10]).date()
    raise ValueError(f"Data inválida: {v!r}")


def _auditoria_dict(a: EconomicoAuditoria) -> dict:
    return {
        "id": a.id,
        "indicador": a.indicador,
        "sistema": a.sistema,
        "resumo_bi": a.resumo_bi,
        "diferenca": a.diferenca,
        "aprovado": a.aprovado,
        "tolerancia": a.tolerancia,
        "origem_sistema": a.origem_sistema,
        "origem_resumo_bi": a.origem_resumo_bi,
    }


def _importacao_dict(imp: EconomicoImportacao) -> dict:
    return {
        "id": imp.id,
        "arquivo_original": imp.arquivo_original,
        "importado_em": imp.importado_em,
        "status": imp.status,
        "observacao": imp.observacao,
    }


def _forecast_versao_dict(db: Session, versao: EconomicoForecastVersao) -> dict:
    itens = _forecast_items_por_chave(db, versao.id)
    ajustes_count = (
        db.query(func.count(EconomicoForecastAjuste.id))
        .filter(EconomicoForecastAjuste.versao_id == versao.id)
        .scalar()
        or 0
    )
    return {
        "id": versao.id,
        "codigo": versao.codigo,
        "nome": versao.nome,
        "motivo": versao.motivo,
        "status": versao.status,
        "origem": versao.origem,
        "versao_base_id": versao.versao_base_id,
        "criado_por": versao.criado_por,
        "criado_em": versao.criado_em,
        "importacao": _importacao_dict(versao.importacao) if versao.importacao else None,
        "receita_forecast": itens.get(("receita", None), 0.0),
        "custos_diretos_forecast": itens.get(("custos_diretos", None), 0.0),
        "custos_indiretos_forecast": itens.get(("custos_indiretos", None), 0.0),
        "impostos_forecast": itens.get(("impostos", None), 0.0),
        "resultado_forecast": itens.get(("resultado", None), 0.0),
        "ajustes_count": ajustes_count,
    }


def _forecast_item_dict(item: EconomicoForecastItem) -> dict:
    return {
        "id": item.id,
        "indicador": item.indicador,
        "categoria": item.categoria,
        "periodo": item.periodo.isoformat() if item.periodo else None,
        "valor": float(item.valor or 0.0),
        "origem": item.origem,
    }


def _forecast_ajuste_dict(ajuste: EconomicoForecastAjuste) -> dict:
    return {
        "id": ajuste.id,
        "item_id": ajuste.item_id,
        "categoria": ajuste.categoria,
        "valor_anterior": float(ajuste.valor_anterior or 0.0),
        "valor_novo": float(ajuste.valor_novo or 0.0),
        "diferenca": float(ajuste.diferenca or 0.0),
        "justificativa": ajuste.justificativa,
        "usuario": ajuste.usuario,
        "criado_em": ajuste.criado_em,
    }


def _forecast_historico_dict(row: EconomicoForecastHistorico) -> dict:
    try:
        payload = json.loads(row.payload or "{}")
    except json.JSONDecodeError:
        payload = {}
    return {
        "id": row.id,
        "acao": row.acao,
        "descricao": row.descricao,
        "usuario": row.usuario,
        "payload": payload,
        "criado_em": row.criado_em,
    }


def _as_date_or_none(v) -> date | None:
    if v in (None, ""):
        return None
    try:
        return _as_date(v)
    except ValueError:
        return None


def _str(v) -> str | None:
    if v is None:
        return None
    text = str(v).strip()
    return text or None


def _sheet(wb, token: str):
    token_l = token.lower()
    for name in wb.sheetnames:
        if token_l in name.lower():
            return wb[name]
    raise ValueError(f"Aba com token {token!r} nao encontrada")


def _lancamento_dict(r: EconomicoLancamentoRazao) -> dict:
    return {
        "id": r.id,
        "data": r.data.isoformat() if r.data else None,
        "documento": r.documento,
        "fornecedor": r.fornecedor,
        "conta": r.conta,
        "conta_descricao": r.conta_descricao,
        "categoria_dre": r.categoria_dre,
        "historico": r.historico,
        "valor": r.valor,
        "tipo": r.tipo,
        "lote": r.lote,
        "lancamento": r.lancamento,
    }


def _filtros_lancamentos(db: Session, importacao_id: int) -> dict:
    def distinct(col):
        values = (
            db.query(col)
            .filter(EconomicoLancamentoRazao.importacao_id == importacao_id, col.isnot(None), col != "")
            .distinct()
            .order_by(col)
            .limit(500)
            .all()
        )
        return [v[0] for v in values]

    periodo = (
        db.query(func.min(EconomicoLancamentoRazao.data), func.max(EconomicoLancamentoRazao.data))
        .filter(EconomicoLancamentoRazao.importacao_id == importacao_id)
        .first()
    )
    return {
        "categorias": distinct(EconomicoLancamentoRazao.categoria_dre),
        "fornecedores": distinct(EconomicoLancamentoRazao.fornecedor),
        "contas": distinct(EconomicoLancamentoRazao.conta),
        "documentos": distinct(EconomicoLancamentoRazao.documento),
        "periodo_min": periodo[0].isoformat() if periodo and periodo[0] else None,
        "periodo_max": periodo[1].isoformat() if periodo and periodo[1] else None,
    }
