from __future__ import annotations

import os
import shutil
from functools import lru_cache
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook


DEFAULT_CATALOGO_PATH = Path(
    r"C:\Users\lucas.barros\Downloads\ET-5275.00-2000-293-E6G-001=0"
    r"\catalogo_tecnico_parametrizado_suportes_v7_regras_corrigido.xlsx"
)

UPLOAD_CATALOGO_PATH = Path(__file__).resolve().parents[1] / "uploads" / "catalogo_suportes.xlsx"
ET_SUPORTES_UPLOAD_URL = "/uploads/et_suportes"

SHEETS = [
    "catalogo_suportes",
    "materiais_suporte",
    "dimensoes_suporte",
    "componentes_referenciados",
    "variaveis_suporte",
    "observacoes_regras",
    "pendencias_validacao",
    "solicitacao_modelo",
]

SHEETS_LIMPEZA_103 = [
    "materiais_suporte",
    "componentes_referenciados",
    "dimensoes_suporte",
    "variaveis_suporte",
    "observacoes_regras",
    "pendencias_validacao",
    "solicitacao_modelo",
]

CAMPOS_103_INVALIDO = {
    "item",
    "tipo",
    "diametro_tubo",
    "faixa_diametro_tubo",
    "isolamento",
    "espessura_isolamento",
    "componente_referenciado",
    "condicao_material",
    "observacao_material",
    "observacao",
    "status_validacao_material",
    "material",
    "dimensao_especificacao",
    "tipo_material",
    "condicional",
    "condicao_aplicacao",
}

CAMPOS_103_VALIDO = {"folha_et"}

FILTROS_MATERIAIS = [
    "codigo_suporte",
    "material",
    "tipo_material",
    "diametro_tubo",
    "faixa_diametro_tubo",
    "tipo",
    "item",
    "condicional",
    "status_validacao_material",
]

CAMPOS_SOLICITACAO_BASE = [
    "solicitante",
    "projeto",
    "linha",
    "codigo_suporte",
    "item",
    "tipo",
    "diametro_tubo",
    "faixa_diametro_tubo",
    "isolamento",
    "espessura_isolamento",
    "estrutura_fixacao",
    "material_linha",
    "quantidade_suportes",
    "observacao_solicitante",
]

SUPLEMENTO_MATERIAIS_VISUAIS: dict[str, list[dict[str, Any]]] = {
    "SA-01": [
        {"tipo": "Tipo II", "faixa_diametro_tubo": '10" até 14"', "material": "Perfil I", "dimensao_especificacao": "1/2 I 203 x 27,3 kg/m", "quantidade_unitaria": "5.5", "unidade": "kg", "condicional": "Não", "condicao_material": "Tipo II conforme folha ET 63; espessura do isolamento até 80 mm."},
        {"tipo": "Tipo II", "faixa_diametro_tubo": '10" até 14"', "material": "Perfil I", "dimensao_especificacao": "1/2 I 203 x 27,3 kg/m", "quantidade_unitaria": "6.5", "unidade": "kg", "condicional": "Não", "condicao_material": "Tipo II conforme folha ET 63; espessura do isolamento de 80 até 120 mm."},
        {"tipo": "Tipo II", "faixa_diametro_tubo": '10" até 14"', "material": "Chapa esp. 6,4", "dimensao_especificacao": "Chapa 6,4", "quantidade_unitaria": "1.0", "unidade": "kg", "condicional": "Não", "condicao_material": "Tipo II conforme folha ET 63; espessura do isolamento até 80 mm."},
        {"tipo": "Tipo II", "faixa_diametro_tubo": '10" até 14"', "material": "Chapa esp. 6,4", "dimensao_especificacao": "Chapa 6,4", "quantidade_unitaria": "1.5", "unidade": "kg", "condicional": "Não", "condicao_material": "Tipo II conforme folha ET 63; espessura do isolamento de 80 até 120 mm."},
    ],
    "BA-01": [
        {"material": "Placa base do suporte", "dimensao_especificacao": "Conforme folha ET; B e D conforme suporte correspondente", "quantidade_unitaria": "CONFIRMAR", "unidade": "pç", "condicional": "Não", "condicao_material": ""},
        {"material": "Chumbador", "dimensao_especificacao": "Conforme folha ET", "quantidade_unitaria": "4", "unidade": "pç", "condicional": "Sim", "condicao_material": "Quando requerido"},
    ],
    "BA-02": [
        {"material": "Inserto", "dimensao_especificacao": "Conforme detalhe do inserto na folha ET", "quantidade_unitaria": "CONFIRMAR", "unidade": "pç", "condicional": "Não", "condicao_material": ""},
        {"material": "Chapa esp. 12,7", "dimensao_especificacao": "Conforme folha ET", "quantidade_unitaria": "CONFIRMAR", "unidade": "kg", "condicional": "Não", "condicao_material": ""},
        {"material": "Barra Ø10 - Aço CA-25", "dimensao_especificacao": "Ø10 - 250", "quantidade_unitaria": "CONFIRMAR", "unidade": "pç", "condicional": "Não", "condicao_material": ""},
    ],
}

SUPLEMENTO_DIMENSOES_VISUAIS: dict[str, list[dict[str, Any]]] = {
    "AC-04": [
        {"tipo": "Tipo I", "diametro_tubo": '1/2"-1 1/2"', "faixa_diametro_tubo": '1/2" a 14"', "carga_maxima_kgf": "50", "momento_kgfm": "20"},
        {"tipo": "Tipo I", "diametro_tubo": '2"-6"', "faixa_diametro_tubo": '1/2" a 14"', "carga_maxima_kgf": "900", "momento_kgfm": "730"},
        {"tipo": "Tipo I", "diametro_tubo": '8"', "faixa_diametro_tubo": '1/2" a 14"', "carga_maxima_kgf": "1430", "momento_kgfm": "1260"},
        {"tipo": "Tipo I", "diametro_tubo": '10"', "faixa_diametro_tubo": '1/2" a 14"', "carga_maxima_kgf": "2340", "momento_kgfm": "2190"},
        {"tipo": "Tipo I", "diametro_tubo": '12"', "faixa_diametro_tubo": '1/2" a 14"', "carga_maxima_kgf": "3640", "momento_kgfm": "3790"},
        {"tipo": "Tipo I", "diametro_tubo": '14"', "faixa_diametro_tubo": '1/2" a 14"', "carga_maxima_kgf": "4460", "momento_kgfm": "4790"},
        {"tipo": "Tipo II", "diametro_tubo": '16"', "faixa_diametro_tubo": '16" a 24"', "carga_maxima_kgf": "6010", "momento_kgfm": "6710"},
        {"tipo": "Tipo II", "diametro_tubo": '18"', "faixa_diametro_tubo": '16" a 24"', "carga_maxima_kgf": "7800", "momento_kgfm": "8970"},
        {"tipo": "Tipo II", "diametro_tubo": '20"', "faixa_diametro_tubo": '16" a 24"', "carga_maxima_kgf": "9810", "momento_kgfm": "11520"},
        {"tipo": "Tipo II", "diametro_tubo": '22"', "faixa_diametro_tubo": '16" a 24"', "carga_maxima_kgf": "10800", "momento_kgfm": "12960"},
        {"tipo": "Tipo II", "diametro_tubo": '24"', "faixa_diametro_tubo": '16" a 24"', "carga_maxima_kgf": "14650", "momento_kgfm": "17950"},
        {"tipo": "Tipo III", "diametro_tubo": '26"-28"', "faixa_diametro_tubo": '26" a 48"', "carga_maxima_kgf": "7800", "momento_kgfm": "9750"},
        {"tipo": "Tipo III", "diametro_tubo": '30"', "faixa_diametro_tubo": '26" a 48"', "carga_maxima_kgf": "10900", "momento_kgfm": "13900"},
        {"tipo": "Tipo III", "diametro_tubo": '32"', "faixa_diametro_tubo": '26" a 48"', "carga_maxima_kgf": "11660", "momento_kgfm": "15350"},
        {"tipo": "Tipo III", "diametro_tubo": '34"', "faixa_diametro_tubo": '26" a 48"', "carga_maxima_kgf": "12500", "momento_kgfm": "16730"},
        {"tipo": "Tipo III", "diametro_tubo": '36"', "faixa_diametro_tubo": '26" a 48"', "carga_maxima_kgf": "13440", "momento_kgfm": "18220"},
        {"tipo": "Tipo III", "diametro_tubo": '38"', "faixa_diametro_tubo": '26" a 48"', "carga_maxima_kgf": "13440", "momento_kgfm": "18220"},
        {"tipo": "Tipo III", "diametro_tubo": '40"', "faixa_diametro_tubo": '26" a 48"', "carga_maxima_kgf": "16600", "momento_kgfm": "23000"},
        {"tipo": "Tipo III", "diametro_tubo": '44"', "faixa_diametro_tubo": '26" a 48"', "carga_maxima_kgf": "20540", "momento_kgfm": "29100"},
        {"tipo": "Tipo III", "diametro_tubo": '48"', "faixa_diametro_tubo": '26" a 48"', "carga_maxima_kgf": "25900", "momento_kgfm": "37350"},
    ]
}

for _codigo_clip in [
    "CL-01A", "CL-01B", "CL-02", "CL-03", "CL-04", "CL-05", "CL-06", "CL-07",
    "CL-08", "CL-09", "CL-10", "CL-11", "CL-12",
]:
    SUPLEMENTO_MATERIAIS_VISUAIS[_codigo_clip] = [
        {
            "material": "Chapa esp. 9,5",
            "dimensao_especificacao": "Conforme folha ET e tipo aplicável",
            "quantidade_unitaria": "CONFIRMAR",
            "unidade": "kg",
            "condicional": "Não",
            "condicao_material": "Material do clip deve ser o mesmo do equipamento ou compatível",
        }
    ]


def _catalogo_path() -> Path:
    env_path = os.getenv("CATALOGO_SUPORTES_XLSX")
    if env_path:
        return Path(env_path)
    if UPLOAD_CATALOGO_PATH.exists():
        return UPLOAD_CATALOGO_PATH
    return DEFAULT_CATALOGO_PATH


def _norm_header(value: Any) -> str:
    return str(value or "").strip()


def _norm_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _is_103(value: Any) -> bool:
    return str(value).strip() == "103"


def _clean_value(sheet: str, campo: str, value: Any, report: dict[str, Any]) -> Any:
    if not _is_103(value):
        return _norm_value(value)

    ocorrencia = {"aba": sheet, "campo": campo, "valor": "103"}
    report["encontradas"] += 1

    if campo in CAMPOS_103_VALIDO:
        ocorrencia["acao"] = "mantido"
        ocorrencia["justificativa"] = "Campo folha/pagina/fonte tecnica valida."
        report["mantidas"] += 1
        report["detalhes"].append(ocorrencia)
        return _norm_value(value)

    if campo in CAMPOS_103_INVALIDO:
        ocorrencia["acao"] = "removido"
        ocorrencia["justificativa"] = "Campo tecnico onde 103 nao representa opcao valida."
        report["removidas"] += 1
        report["detalhes"].append(ocorrencia)
        return ""

    ocorrencia["acao"] = "mantido"
    ocorrencia["justificativa"] = "Campo nao listado como tecnico invalido para 103."
    report["mantidas"] += 1
    report["detalhes"].append(ocorrencia)
    return _norm_value(value)


def _empty_report() -> dict[str, Any]:
    return {
        "encontradas": 0,
        "removidas": 0,
        "mantidas": 0,
        "detalhes": [],
        "regra": {
            "remover_em": sorted(CAMPOS_103_INVALIDO),
            "manter_em": sorted(CAMPOS_103_VALIDO),
        },
    }


def _read_sheet(wb, sheet: str, report: dict[str, Any]) -> list[dict[str, Any]]:
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    headers = [_norm_header(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v not in (None, "") for v in values):
            continue
        item: dict[str, Any] = {}
        for header, value in zip(headers, values):
            if not header:
                continue
            if sheet in SHEETS_LIMPEZA_103:
                item[header] = _clean_value(sheet, header, value, report)
            else:
                item[header] = _norm_value(value)
        item["_linha_excel"] = excel_row
        rows.append(item)
    return rows


def _mtime(path: Path) -> float:
    return path.stat().st_mtime if path.exists() else 0.0


@lru_cache(maxsize=4)
def _load_catalogo_cached(path_str: str, modified: float) -> dict[str, Any]:
    path = Path(path_str)
    if not path.exists():
        return {"erro": f"Planilha nao encontrada: {path}", "abas": {}, "limpeza_103": _empty_report()}

    report = _empty_report()
    wb = load_workbook(path, read_only=True, data_only=True)
    abas = {sheet: _read_sheet(wb, sheet, report) for sheet in SHEETS}
    wb.close()
    return {
        "arquivo": str(path),
        "modificado_em": modified,
        "abas": abas,
        "limpeza_103": report,
    }


def carregar_catalogo() -> dict[str, Any]:
    path = _catalogo_path()
    return _load_catalogo_cached(str(path), _mtime(path))


def salvar_catalogo_upload(file: UploadFile) -> dict[str, Any]:
    UPLOAD_CATALOGO_PATH.parent.mkdir(parents=True, exist_ok=True)
    with UPLOAD_CATALOGO_PATH.open("wb") as dest:
        shutil.copyfileobj(file.file, dest)
    _load_catalogo_cached.cache_clear()
    return resumo_catalogo()


def _eq_filter(value: Any, expected: str) -> bool:
    if expected in (None, ""):
        return True
    return str(value or "").strip().lower() == str(expected).strip().lower()


def _contains_filter(value: Any, expected: str) -> bool:
    if expected in (None, ""):
        return True
    return str(expected).strip().lower() in str(value or "").strip().lower()


def _norm_search(value: Any) -> str:
    return str(value or "").strip().lower().replace('"', "").replace("”", "").replace("″", "")


def _dn_number(value: Any) -> float | None:
    texto = _norm_search(value)
    texto = texto.replace("até", "a").replace("ate", "a").replace("de ", "").strip()
    if not texto:
        return None
    import re
    misto = re.match(r"^(\d+(?:\.\d+)?)\s+(\d+)\s*/\s*(\d+)", texto)
    if misto:
        try:
            inteiro, n, d = misto.groups()
            return float(inteiro) + float(n) / float(d)
        except (ValueError, ZeroDivisionError):
            return None
    token = texto.split()[0].replace(",", ".")
    if "-" in token:
        token = token.split("-", 1)[0]
    partes = token.split()
    if len(partes) == 2 and "/" in partes[1]:
        try:
            n, d = partes[1].split("/", 1)
            return float(partes[0]) + float(n) / float(d)
        except (ValueError, ZeroDivisionError):
            return None
    if "/" in token:
        try:
            n, d = token.split("/", 1)
            return float(n) / float(d)
        except (ValueError, ZeroDivisionError):
            return None
    try:
        return float(token)
    except ValueError:
        return None


def _dn_interval(value: Any) -> tuple[float, float] | None:
    texto = _norm_search(value)
    texto = texto.replace("até", "a").replace("ate", "a").replace("de ", "")
    if " a " in texto:
        ini, fim = texto.split(" a ", 1)
    elif "-" in texto:
        ini, fim = texto.split("-", 1)
    else:
        dn = _dn_number(texto)
        return (dn, dn) if dn is not None else None
    dn_ini = _dn_number(ini)
    dn_fim = _dn_number(fim)
    if dn_ini is None or dn_fim is None:
        return None
    return (min(dn_ini, dn_fim), max(dn_ini, dn_fim))


def _dn_matches_value(value: Any, expected: Any) -> bool:
    alvo = _dn_number(expected)
    intervalo = _dn_interval(value)
    if alvo is None or intervalo is None:
        return False
    return intervalo[0] <= alvo <= intervalo[1]


def _is_wildcard_value(value: Any) -> bool:
    texto = _norm_search(value)
    return texto in {"", "todos", "todas", "geral", "aplicavel a todos", "aplicável a todos"}


def _match_select_value(row_value: Any, expected: str | None) -> bool:
    if expected in (None, ""):
        return True
    if _is_wildcard_value(row_value):
        return True
    alvo = _norm_search(expected)
    valor = _norm_search(row_value)
    if alvo == valor:
        return True
    partes = [p.strip() for p in valor.replace(" ou ", "/").split("/") if p.strip()]
    return alvo in partes


def _match_diametro(row: dict[str, Any], dn: str | None) -> bool:
    if dn in (None, ""):
        return True
    alvo = _norm_search(dn)
    diametro = row.get("diametro_tubo")
    faixa = row.get("faixa_diametro_tubo")
    if _is_wildcard_value(diametro) and _is_wildcard_value(faixa):
        return True
    valores = [diametro] if not _is_wildcard_value(diametro) else [faixa, row.get("tipo")]
    for valor in valores:
        texto = _norm_search(valor)
        if not texto:
            continue
        if texto == alvo:
            return True
        if _dn_matches_value(valor, dn):
            return True
        partes = [p.strip() for p in texto.replace(" ou ", "/").split("/") if p.strip()]
        if alvo in partes:
            return True
    return False


def _folha_links(folha_et: Any) -> dict[str, Any] | None:
    folha = str(folha_et or "").strip()
    if not folha:
        return None
    try:
        numero = int(float(folha.replace(",", ".")))
    except ValueError:
        return None
    if numero <= 0:
        return None
    return {
        "folha": str(numero),
        "png_url": f"{ET_SUPORTES_UPLOAD_URL}/png/folha_{numero:03d}.png",
        "pdf_url": f"{ET_SUPORTES_UPLOAD_URL}/pdf/folha_{numero:03d}.pdf",
    }


def _materiais_suplementares(codigo: str, folha_et: Any) -> list[dict[str, Any]]:
    linhas = []
    for idx, row in enumerate(SUPLEMENTO_MATERIAIS_VISUAIS.get(codigo.upper(), []), start=1):
        linhas.append({
            "codigo_suporte": codigo.upper(),
            "folha_et": folha_et or "",
            "item": "",
            "tipo": row.get("tipo", ""),
            "diametro_tubo": row.get("diametro_tubo", ""),
            "faixa_diametro_tubo": row.get("faixa_diametro_tubo", ""),
            "material": row.get("material", ""),
            "dimensao_especificacao": row.get("dimensao_especificacao", ""),
            "quantidade_unitaria": row.get("quantidade_unitaria", "CONFIRMAR"),
            "unidade": row.get("unidade", ""),
            "tipo_material": "Direto",
            "componente_referenciado": "",
            "explodir_componente": "Não",
            "condicional": row.get("condicional", "Não"),
            "condicao_material": row.get("condicao_material", ""),
            "status_validacao_material": "Transcrição parcial visual",
            "observacao_material": (
                "Incluído por auditoria visual da folha ET. "
                "Quantidade/detalhamento deve ser confirmado antes de compra/MTO."
            ),
            "isolamento": row.get("isolamento", ""),
            "espessura_isolamento": row.get("espessura_isolamento", ""),
            "_linha_excel": f"suplemento-{idx}",
        })
    return linhas


def _dimensoes_suplementares(codigo: str, folha_et: Any) -> list[dict[str, Any]]:
    linhas = []
    for idx, row in enumerate(SUPLEMENTO_DIMENSOES_VISUAIS.get(codigo.upper(), []), start=1):
        linhas.append({
            "codigo_suporte": codigo.upper(),
            "folha_et": folha_et or "",
            "item": "",
            "tipo": row.get("tipo", ""),
            "diametro_tubo": row.get("diametro_tubo", ""),
            "faixa_diametro_tubo": row.get("faixa_diametro_tubo", ""),
            "isolamento": "",
            "espessura_isolamento": "",
            "tipo_aplicacao": "",
            "carga_maxima_kgf": row.get("carga_maxima_kgf", ""),
            "carga_recomendada_kgf": "",
            "forca_kgf": "",
            "momento_kgfm": row.get("momento_kgfm", ""),
            "A": "",
            "B": "",
            "B1": "",
            "C": "",
            "D": "",
            "E": "",
            "F": "",
            "G": "",
            "H": "",
            "L": "",
            "M": "",
            "N": "",
            "R": "",
            "t": "",
            "d1": "",
            "d2": "",
            "X": "",
            "J": "",
            "dimensao_variavel": "DN detalhado conforme tabela da folha ET",
            "fonte_dimensao_variavel": "ET PDF folha 8",
            "status_validacao_dimensional": "Validado visualmente na ET PDF",
            "observacao_dimensional": "Suplemento visual para seleção operacional do DN; não altera a planilha base.",
            "_linha_excel": f"suplemento-dim-{idx}",
        })
    return linhas


def resumo_catalogo() -> dict[str, Any]:
    data = carregar_catalogo()
    abas = data.get("abas", {})
    codigos = sorted({str(r.get("codigo_suporte", "")).strip() for r in abas.get("catalogo_suportes", []) if r.get("codigo_suporte")})
    return {
        "arquivo": data.get("arquivo"),
        "totais": {sheet: len(rows) for sheet, rows in abas.items()},
        "codigos_suporte": codigos,
        "limpeza_103": data.get("limpeza_103"),
    }


def listar_materiais(filtros: dict[str, str], limite: int = 500) -> dict[str, Any]:
    data = carregar_catalogo()
    rows = data.get("abas", {}).get("materiais_suporte", [])
    out = []
    for row in rows:
        ok = True
        for campo in FILTROS_MATERIAIS:
            esperado = filtros.get(campo)
            if campo in {"codigo_suporte", "material", "dimensao_especificacao"}:
                ok = ok and _contains_filter(row.get(campo), esperado)
            else:
                ok = ok and _eq_filter(row.get(campo), esperado)
        if ok:
            out.append(row)
        if len(out) >= limite:
            break
    return {"total": len(out), "limite": limite, "materiais": out}


def _match_operacional(row: dict[str, Any], item: str | None, tipo: str | None, dn: str | None) -> bool:
    if item and not _match_select_value(row.get("item"), item):
        return False
    if tipo and not _match_select_value(row.get("tipo"), tipo):
        return False
    if not _match_diametro(row, dn):
        return False
    return True


def consulta_operacional(
    codigo_suporte: str,
    item: str | None = None,
    tipo: str | None = None,
    dn: str | None = None,
) -> dict[str, Any] | None:
    detalhe = detalhe_suporte(codigo_suporte)
    if not detalhe:
        return None

    materiais = [
        r for r in detalhe["materiais_diretos"] + detalhe["materiais_componentes"]
        if _match_operacional(r, item, tipo, dn)
    ]
    componentes = [
        r for r in detalhe["componentes_referenciados"]
        if _match_operacional(r, item, tipo, dn)
    ]
    dimensoes = [
        r for r in detalhe["dimensoes"]
        if _match_operacional(r, item, tipo, dn)
    ]
    variaveis = detalhe["variaveis"]

    aviso_componentes = None
    if not materiais and componentes:
        aviso_componentes = "Este suporte e composto por componentes referenciados. Consulte a tabela de componentes abaixo."
    elif not materiais and not componentes:
        aviso_componentes = "Nao ha materiais diretos ou componentes referenciados cadastrados para este suporte no catalogo carregado."

    return {
        "codigo_suporte": detalhe["codigo_suporte"],
        "catalogo": detalhe["catalogo"],
        "folha_referencia": _folha_links((detalhe["catalogo"] or {}).get("folha_et")),
        "materiais": materiais,
        "componentes_referenciados": componentes,
        "dimensoes_suporte": dimensoes,
        "variaveis_suporte": variaveis,
        "aviso_componentes": aviso_componentes,
    }


def opcoes_filtros() -> dict[str, list[Any]]:
    abas = carregar_catalogo().get("abas", {})
    rows = abas.get("materiais_suporte", [])
    opts: dict[str, set[Any]] = {campo: set() for campo in FILTROS_MATERIAIS}
    for row in abas.get("catalogo_suportes", []):
        codigo = row.get("codigo_suporte")
        if codigo:
            opts["codigo_suporte"].add(codigo)
    for row in rows:
        for campo in FILTROS_MATERIAIS:
            val = row.get(campo)
            if val not in ("", None) and not (campo in CAMPOS_103_INVALIDO and _is_103(val)):
                opts[campo].add(val)
    return {campo: sorted(vals, key=lambda x: str(x)) for campo, vals in opts.items()}


def detalhe_suporte(codigo: str) -> dict[str, Any] | None:
    data = carregar_catalogo()
    abas = data.get("abas", {})
    codigo_norm = codigo.strip().upper()

    def by_codigo(sheet: str) -> list[dict[str, Any]]:
        return [r for r in abas.get(sheet, []) if str(r.get("codigo_suporte", "")).strip().upper() == codigo_norm]

    catalogo = by_codigo("catalogo_suportes")
    materiais = by_codigo("materiais_suporte")
    dimensoes = by_codigo("dimensoes_suporte")
    componentes = by_codigo("componentes_referenciados")
    variaveis = by_codigo("variaveis_suporte")
    regras = by_codigo("observacoes_regras")
    pendencias = by_codigo("pendencias_validacao")

    if not any([catalogo, materiais, dimensoes, componentes, variaveis, regras, pendencias]):
        return None

    folha_ref = (catalogo[0] if catalogo else {}).get("folha_et", "")
    materiais = materiais + _materiais_suplementares(codigo_norm, folha_ref)
    dimensoes = dimensoes + _dimensoes_suplementares(codigo_norm, folha_ref)

    materiais_diretos = [
        r for r in materiais
        if str(r.get("tipo_material", "")).strip().lower() != "componente referenciado"
    ]
    materiais_componentes = [
        r for r in materiais
        if str(r.get("tipo_material", "")).strip().lower() == "componente referenciado"
    ]
    variaveis_nomes = [str(v.get("variavel") or "").strip() for v in variaveis if v.get("variavel")]
    campos_solicitacao = []
    for campo in CAMPOS_SOLICITACAO_BASE + variaveis_nomes:
        if campo and campo not in campos_solicitacao:
            campos_solicitacao.append(campo)

    return {
        "codigo_suporte": codigo_norm,
        "catalogo": catalogo[0] if catalogo else None,
        "materiais_diretos": materiais_diretos,
        "materiais_componentes": materiais_componentes,
        "dimensoes": dimensoes,
        "componentes_referenciados": componentes,
        "variaveis": variaveis,
        "regras": regras,
        "pendencias": pendencias,
        "campos_solicitacao": campos_solicitacao,
        "avisos": _avisos(materiais, dimensoes, componentes, pendencias),
    }


def _avisos(
    materiais: list[dict[str, Any]],
    dimensoes: list[dict[str, Any]],
    componentes: list[dict[str, Any]],
    pendencias: list[dict[str, Any]],
) -> list[str]:
    avisos: list[str] = []
    if pendencias:
        avisos.append("Suporte possui pendencias de validacao.")
    if any(str(r.get("quantidade_unitaria", "")).strip().upper() == "CONFIRMAR" for r in materiais):
        avisos.append("Ha materiais com quantidade unitaria CONFIRMAR.")
    if any(str(r.get("condicional", "")).strip().lower() == "sim" for r in materiais + componentes):
        avisos.append("Ha itens condicionais; validar condicao antes de solicitar.")
    if any(str(r.get("dimensao_variavel", "")).strip().lower() == "sim" for r in dimensoes):
        avisos.append("Ha dimensoes variaveis; informar valores na solicitacao.")
    if componentes:
        avisos.append("Componentes referenciados devem ser consultados separadamente; nao ha explosao automatica.")
    return avisos
