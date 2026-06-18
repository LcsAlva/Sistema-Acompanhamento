"""
Parser XLSX para o Sistema de Programação Semanal URFCC.

Suporta dois formatos:
  1. Exportação ETM interna: abas 'Programacao Folha 01/02'
  2. Exportação nativa Primavera P6: aba 'TASK' com 2 linhas de cabeçalho
     (row 1 = nomes técnicos internos, row 2 = descrições em português)

Mapeamento de colunas por cabeçalho (case-insensitive, strip):
  activity_id  -> "task_code" / "Activity ID" / "ID da Atividade"
  nome         -> "task_name" / "Activity Name" / "Nome da Atividade"
  area_unidade -> "wbs_name" / "Nome da WBS" / "Área" / "Area"
  disciplina   -> "actv_code_urfcc_fase_id" / "Disciplina" / "URFCC-Fase"
  duracao      -> "target_drtn_hr_cnt" / "Original Duration"
  pct_avanco   -> "sched_complete_pct" / "% Complete"
  inicio_prog  -> "start_date" / "BL Project Start" / "Early Start"
  termino_prog -> "end_date" / "BL Project Finish" / "Early Finish"
  inicio_lb    -> "base_start_date" / "Start" / "Início LB"
  termino_lb   -> "base_end_date" / "Finish" / "Término LB"
"""

import openpyxl
import re
from datetime import date, datetime
from typing import Optional, IO

DISCIPLINA_MAP = {
    "0": "Marcos",
    "1": "Mobilização",
    "2": "Engenharia de detalhamento",
    "3": "Construção Civil",
    "4": "Eletromecânica",
    "5": "Comissionamento",
    "6": "Fornecimento de bens",
}

_WBS_RE = re.compile(r'URFCC-\d{4}-\d{2}-\d{2}\.(\d)')

def _disciplina_do_wbs(wbs_codigo: Optional[str]) -> Optional[str]:
    """Extrai o nome da disciplina a partir do código WBS."""
    if not wbs_codigo:
        return None
    m = _WBS_RE.search(str(wbs_codigo))
    if m:
        return DISCIPLINA_MAP.get(m.group(1))
    return None


# Mapeamento cabeçalho -> campo interno
COLUMN_MAP = {
    # activity_id — Primavera interno + ETM
    "task_code": "activity_id",
    "activity id": "activity_id",
    "id da atividade": "activity_id",
    "id": "activity_id",
    "codigo": "activity_id",
    "cód.": "activity_id",
    "cod.": "activity_id",
    # nome
    "task_name": "nome",
    "nome da atividade": "nome",
    "activity name": "nome",
    "nome": "nome",
    "descrição": "nome",
    "descricao": "nome",
    # wbs_codigo — código WBS (ex: URFCC-2026-04-05.1.2.1.4)
    "wbs_id": "wbs_codigo",
    "código da wbs": "wbs_codigo",
    "codigo da wbs": "wbs_codigo",
    # status_atividade — Status da Atividade (P6: status_code)
    "status_code": "status_atividade",
    "status da atividade": "status_atividade",
    # area_unidade — WBS name como área/unidade
    "wbs_name": "area_unidade",
    "nome da wbs": "area_unidade",
    "área": "area_unidade",
    "area": "area_unidade",
    "unidade": "area_unidade",
    "area/unidade": "area_unidade",
    # disciplina — nome textual (não o ID numérico do Primavera)
    "urfcc-fase": "disciplina",
    "disciplina": "disciplina",
    "disc.": "disciplina",
    # actv_code_urfcc_fase_id é ID numérico no P6 — ignorado intencionalmente
    # supervisor / encarregado
    "supervisor": "supervisor",
    "encarregado": "encarregado",
    # duracao
    "target_drtn_hr_cnt": "duracao",
    "original duration": "duracao",
    "duração original(d)": "duracao",
    "duração": "duracao",
    "duracao": "duracao",
    "dur": "duracao",
    # pct_avanco — PREVISTO: Programar % Concluída (sched_complete_pct)
    "sched_complete_pct": "pct_avanco",
    "programar % concluída(%)": "pct_avanco",
    "(*)programar % concluída(%)": "pct_avanco",
    "% complete": "pct_avanco",
    "% avanco": "pct_avanco",
    "% avanço": "pct_avanco",
    "avanco": "pct_avanco",
    "avanço": "pct_avanco",
    # pct_executado — EXECUTADO: % Concluída de Unidades Não Relacionadas à MdO (equip_complete_pct)
    "equip_complete_pct": "pct_executado",
    "(*)% concluída de unidades não relacionadas à mão de obra(%)": "pct_executado",
    "(*)% concluida de unidades nao relacionadas a mao de obra(%)": "pct_executado",
    "% concluída de unidades não relacionadas à mão de obra": "pct_executado",
    "% concluida de unidades nao relacionadas a mao de obra": "pct_executado",
    # unid_orcadas_smo — Unidades Orçadas Sem Mão de Obra (P6: target_equip_qty)
    "target_equip_qty": "unid_orcadas_smo",
    "target_non_labor_qty": "unid_orcadas_smo",
    "unidades orçadas sem mão de obra": "unid_orcadas_smo",
    "unidades orcadas sem mao de obra": "unid_orcadas_smo",
    "(*)unidades orçadas sem mão de obra(h)": "unid_orcadas_smo",
    "(*)unidades orcadas sem mao de obra(h)": "unid_orcadas_smo",
    "unid. orçadas s/ mdo": "unid_orcadas_smo",
    "unid. orcadas s/ mdo": "unid_orcadas_smo",
    "unid orçadas s/ mdo": "unid_orcadas_smo",
    "unid orcadas s/ mdo": "unid_orcadas_smo",
    "unid. orçadas s/mdo": "unid_orcadas_smo",
    "unidades orçadas s/ mão de obra": "unid_orcadas_smo",
    "unidades orçadas s/mão de obra": "unid_orcadas_smo",
    # inicio_prog — data de início programada (Early Start / start_date P6)
    "start_date": "inicio_prog",
    "bl project start": "inicio_prog",
    "início prog": "inicio_prog",
    "inicio prog": "inicio_prog",
    "early start": "inicio_prog",
    # termino_prog — data de término programada
    "end_date": "termino_prog",
    "bl project finish": "termino_prog",
    "término prog": "termino_prog",
    "termino prog": "termino_prog",
    "early finish": "termino_prog",
    # inicio_lb — linha de base
    "base_start_date": "inicio_lb",
    "start": "inicio_lb",
    "início lb": "inicio_lb",
    "inicio lb": "inicio_lb",
    "data início": "inicio_lb",
    "data inicio": "inicio_lb",
    # termino_lb — linha de base
    "base_end_date": "termino_lb",
    "finish": "termino_lb",
    "término lb": "termino_lb",
    "termino lb": "termino_lb",
    "data término": "termino_lb",
    "data termino": "termino_lb",
}

ABAS_PREFERIDAS = [
    "TASK",  # Primavera P6 nativo
    "Programacao Folha 01", "Programacao Folha 02",
    "Programação Folha 01", "Programação Folha 02",
]


def parse_xlsx(file_obj: IO[bytes], aba_preferida: Optional[str] = None) -> list[dict]:
    """
    Lê um arquivo XLSX e retorna lista de dicts com os campos das tarefas.
    Suporta exportação Primavera P6 nativa (aba TASK com 2 cabeçalhos) e ETM.
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)

    aba = _selecionar_aba(wb, aba_preferida)
    ws = wb[aba]

    # Detecta linha de cabeçalho
    header_row, col_map = _detectar_cabecalho(ws)
    if not col_map:
        abas_disponiveis = ", ".join(f"'{s}'" for s in wb.sheetnames)
        raise ValueError(
            f"Não foi possível mapear colunas na aba '{aba}'. "
            f"Abas disponíveis no arquivo: {abas_disponiveis}. "
            "Informe o nome correto da aba no campo 'ABA DO ARQUIVO'."
        )

    # Exportação Primavera tem row 2 como subtítulo em português — detecta e pula
    data_start = header_row + 1
    subtitle_row = next(ws.iter_rows(min_row=data_start, max_row=data_start, values_only=True), None)
    if subtitle_row and _e_linha_subtitulo(subtitle_row, col_map):
        data_start += 1

    tarefas = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if all(v is None for v in row):
            continue

        tarefa = {}
        for col_idx, campo in col_map.items():
            valor = row[col_idx] if col_idx < len(row) else None
            tarefa[campo] = _converter_valor(campo, valor)

        # Pula linhas sem activity_id válido (IDs P6 não têm espaços)
        aid = tarefa.get("activity_id")
        if not aid or " " in str(aid):
            continue

        # Deriva disciplina do código WBS se não vier mapeada diretamente
        if not tarefa.get("disciplina"):
            tarefa["disciplina"] = _disciplina_do_wbs(tarefa.get("wbs_codigo"))

        tarefas.append(tarefa)

    return tarefas


def _e_linha_subtitulo(row: tuple, col_map: dict) -> bool:
    """
    Retorna True se a linha parece ser uma linha de subtítulo/descrição
    (ex: row 2 do Primavera com nomes em português).
    Critério: activity_id mapeado contém espaço ou é uma string descritiva.
    """
    for col_idx, campo in col_map.items():
        if campo == "activity_id" and col_idx < len(row):
            val = row[col_idx]
            if val and " " in str(val):
                return True
    return False


def _selecionar_aba(wb: openpyxl.Workbook, aba_preferida: Optional[str]) -> str:
    """Seleciona a aba correta do workbook."""
    if aba_preferida and aba_preferida in wb.sheetnames:
        return aba_preferida

    for nome in ABAS_PREFERIDAS:
        if nome in wb.sheetnames:
            return nome

    # Fallback: primeira aba com colunas mapeáveis
    for nome in wb.sheetnames:
        ws = wb[nome]
        _, col_map = _detectar_cabecalho(ws)
        if col_map:
            return nome

    # Último recurso: primeira aba (erro detalhado será gerado depois)
    return wb.sheetnames[0]


def _detectar_cabecalho(ws) -> tuple[int, dict]:
    """
    Percorre as primeiras 10 linhas procurando a linha de cabeçalho.
    Retorna (linha_index_1based, {col_index: campo_interno}).
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        col_map = {}
        for col_idx, cell_val in enumerate(row):
            if cell_val is None:
                continue
            key = str(cell_val).strip().lower()
            if key in COLUMN_MAP:
                col_map[col_idx] = COLUMN_MAP[key]
        if len(col_map) >= 2:  # pelo menos activity_id + nome
            return row_idx, col_map
    return 1, {}


def _converter_valor(campo: str, valor):
    """Converte valor bruto da célula para o tipo correto do campo."""
    if valor is None:
        return None

    if campo in ("inicio_lb", "termino_lb", "inicio_prog", "termino_prog"):
        return _para_date(valor)

    if campo in ("pct_avanco", "pct_executado"):
        try:
            v = float(valor)
            # Primavera exporta como percentual inteiro (100) ou fração (0.85)
            return v * 100 if v <= 1.0 else v
        except (ValueError, TypeError):
            return 0.0

    if campo == "duracao":
        try:
            return int(float(valor))
        except (ValueError, TypeError):
            return None

    if campo == "unid_orcadas_smo":
        try:
            return float(valor)
        except (ValueError, TypeError):
            # Célula vazia ou inválida = 0 (sem unidades orçadas)
            return 0.0

    return str(valor).strip() if valor is not None else None


def _para_date(valor) -> Optional[date]:
    """Converte célula de data para objeto date."""
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        # Primavera: "2025-08-11 08:00:00"
        valor = valor.strip().split(" ")[0]
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(valor, fmt).date()
            except ValueError:
                continue
    return None


# ── Filtro QCRON ────────────────────────────────────────────────────────────

def calcular_qcron(
    tarefas_com_raw: list[tuple],
    semana_inicio: date,
    semana_fim: date,
) -> list[tuple]:
    """
    Aplica o filtro QCRON exato do Primavera P6:
      - inicio_prog <= semana_fim
      - termino_prog >= semana_inicio
      - pct_avanco < 100

    Recebe lista de (Tarefa ORM, dict_raw) e retorna apenas as que passam no filtro.
    """
    resultado = []
    for tarefa, t in tarefas_com_raw:
        inicio = t.get("inicio_prog")
        termino = t.get("termino_prog")
        unid_smo = t.get("unid_orcadas_smo")  # None = coluna ausente no arquivo

        # Usa pct_executado (equip_complete_pct) para exclusão de concluídas.
        # Fallback para pct_avanco se pct_executado não estiver no arquivo.
        pct_exec = t.get("pct_executado")
        pct_prev = t.get("pct_avanco", 0.0) or 0.0
        pct_filtro = pct_exec if pct_exec is not None else pct_prev

        if inicio is None or termino is None:
            continue
        # Exclui atividades 100% executadas
        if pct_filtro >= 100:
            continue
        # Exclui atividades com Unidades Orçadas Sem Mão de Obra = 0 ou célula vazia
        # (unid_smo é None apenas quando a coluna não existe no arquivo — aí não filtra)
        if unid_smo is not None and unid_smo <= 0:
            continue
        if inicio <= semana_fim and termino >= semana_inicio:
            resultado.append((tarefa, t))

    return resultado
