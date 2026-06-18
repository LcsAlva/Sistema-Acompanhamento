"""Parser do XLSX da EAP financeira (revisão Petrobras).

Estrutura esperada (aba "EAP"):
  - L4: cabeçalho "NÍVEL", "ITEM", "NÍVEL EAP", "DESCRIÇÃO", "PONDERAÇÃO (R$)"
  - L5: sub-header com colunas Entrega/Fase/.../Detalhamento 2 (cols A..H),
        ITEM (I), NÍVEL EAP (J), ESCOPO (K), VALOR (R$) (L), pesos por nível
        (M..T), % Acumulado (U), R$ Acumulado (V), e a partir de W as
        DATAS MENSAIS (datetime no header).
  - L8 em diante: dados.

Cada item-folha terá:
  codigo (col I), descricao (col K), nivel (col J), valor (col L) e
  dist_mensal {iso_date: fração} extraída das cols mensais.

A função `parse_eap_xlsx(file_obj)` retorna:
  (lista_itens, lista_meses)

onde lista_itens é uma lista de dicts pronta para inserção no banco e
lista_meses são as chaves ISO (yyyy-mm-dd) das colunas mensais detectadas.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from typing import Iterable

import openpyxl


# Cabeçalhos esperados em L5
COL_ITEM = 9         # I
COL_NIVEL = 10       # J
COL_DESC = 11        # K
COL_VALOR = 12       # L
COL_PRIMEIRO_MES = 23  # W (1-indexed)


def _parent_codigo(codigo: str) -> str | None:
    """De '1.2.1.3' devolve '1.2.1'. Para '1' devolve None."""
    if not codigo or '.' not in codigo:
        return None
    return '.'.join(codigo.split('.')[:-1])


def _normalizar_codigo(v) -> str | None:
    """Normaliza o código da EAP. Retorna None se for cabeçalho/vazio."""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # Aceita "1", "1.2", "1.2.1.3" — recusa textos como "ITEM" ou "RAZÃO SOCIAL:"
    parts = s.split('.')
    if not all(p.isdigit() for p in parts):
        return None
    return s


def _to_float(v) -> float:
    if v is None or v == '':
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _to_iso_mes(v) -> str | None:
    """Converte cabeçalho de coluna mensal em 'yyyy-mm-01'.

    Aceita datetime, date ou string 'yyyy-mm-dd hh:mm:ss'.
    """
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.replace(day=1).date().isoformat()
    if isinstance(v, date):
        return v.replace(day=1).isoformat()
    s = str(v).strip()
    # tenta parsear "2025-08-01 00:00:00"
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            d = datetime.strptime(s.split(' ')[0] if ' ' in s else s, fmt.split(' ')[0])
            return d.replace(day=1).date().isoformat()
        except ValueError:
            continue
    return None


def parse_eap_xlsx(file_obj) -> tuple[list[dict], list[str]]:
    """Lê o XLSX da EAP financeira e devolve itens + lista de meses.

    Args:
        file_obj: BytesIO ou caminho para o arquivo .xlsx
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    if 'EAP' not in wb.sheetnames:
        raise ValueError("Aba 'EAP' não encontrada no arquivo.")
    ws = wb['EAP']

    # Extrai os meses do cabeçalho (linha 5, a partir da col W)
    header_l5 = next(ws.iter_rows(min_row=5, max_row=5, values_only=True))
    meses = []
    col_to_mes: dict[int, str] = {}
    for c_idx, valor in enumerate(header_l5, start=1):
        if c_idx < COL_PRIMEIRO_MES:
            continue
        iso = _to_iso_mes(valor)
        if iso:
            col_to_mes[c_idx] = iso
            meses.append(iso)

    itens: list[dict] = []
    seen_codigos: set[str] = set()

    for row in ws.iter_rows(min_row=8, values_only=True):
        if not row:
            continue

        codigo_raw = row[COL_ITEM - 1]
        nivel_raw = row[COL_NIVEL - 1] if len(row) > COL_NIVEL - 1 else None

        # Caso especial: a entrega 1 da EAP da Petrobras vem com o ITEM
        # gravado como "ETM ENGENHARIA" (nome da empresa) em vez de "1".
        # Detectamos via nivel == 1 e adotamos "1" como código.
        if (codigo_raw and not _normalizar_codigo(codigo_raw)
                and nivel_raw == 1
                and 'ETM' in str(codigo_raw).upper()):
            codigo = '1'
        else:
            codigo = _normalizar_codigo(codigo_raw)

        if not codigo or codigo in seen_codigos:
            continue
        seen_codigos.add(codigo)

        descricao_raw = row[COL_DESC - 1]
        if not descricao_raw:
            continue
        descricao = str(descricao_raw).strip()

        nivel_raw = row[COL_NIVEL - 1]
        try:
            nivel = int(nivel_raw) if nivel_raw is not None else len(codigo.split('.'))
        except (TypeError, ValueError):
            nivel = len(codigo.split('.'))

        valor = _to_float(row[COL_VALOR - 1])

        # Distribuição mensal — a planilha mistura formatos:
        #   - frações (0.05, 0.10, soma ~ 1.0) → multiplicar por valor
        #   - valores absolutos em R$ (soma ~ valor da linha) → manter
        # Normalizamos sempre para R$ absoluto.
        dist_raw = {}
        for c_idx, iso in col_to_mes.items():
            if c_idx - 1 < len(row):
                v = _to_float(row[c_idx - 1])
                if v:
                    dist_raw[iso] = v

        soma = sum(dist_raw.values())
        dist = {}
        if dist_raw and valor > 0:
            # Heurística: se soma ≈ 1, está em fração; caso contrário, R$ absoluto.
            # Limite folgado para tolerar arredondamentos.
            if 0.5 <= soma <= 1.5 and abs(soma - valor) > 1.0:
                # fração — converte para R$
                for k, v in dist_raw.items():
                    dist[k] = round(v * valor, 4)
            else:
                # já em R$ absoluto
                dist = {k: round(v, 4) for k, v in dist_raw.items()}
        else:
            dist = dist_raw

        itens.append({
            'codigo': codigo,
            'descricao': descricao,
            'nivel': nivel,
            'parent_codigo': _parent_codigo(codigo),
            'valor': valor,
            'dist_mensal': dist,
        })

    return itens, meses
