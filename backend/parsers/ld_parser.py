"""Parser da Lista de Documentos (LD) recebida da S5 (Módulo 1 — Fase 2A).

Lê o Excel da LD (ex.: LD-5275.00-2000-940-E6G-001) com openpyxl e devolve uma
lista de dicts normalizados, um por documento.

Detecção AUTOMÁTICA de colunas por sinônimos (case-insensitive, strip). O mapa
`COLUMN_MAP` é o ponto de CALIBRAÇÃO: ao receber a amostra real da LD, basta
acrescentar os cabeçalhos exatos do arquivo aqui — nenhuma outra mudança.

Campos de saída (chaves do dict):
  codigo_documento*, titulo, disciplina, revisao, status, a4_equivalente,
  data_prevista, data_emissao
(* obrigatório; linhas sem código são ignoradas)
"""
from __future__ import annotations

import re
from datetime import date, datetime
from typing import IO, Optional

import openpyxl


# Sinônimo de cabeçalho (normalizado) → campo interno.
# CALIBRADO com o layout real LD-5275.00-2000-940-E6G-001 (S5):
#   col2  "NÚMERO N-1710"             → codigo_documento
#   col3  "REVISÃO"                   → revisao
#   col6  "TÍTULO"                    → titulo
#   col8  "DISCIPLINA"                → disciplina
#   col12 "A4 EQUIVALENTE"            → a4_equivalente
#   col14 "DATA PREVISTA DE EMISSÃO"  → data_prevista
#   col16 "DATA EFETIVA DE EMISSÃO"   → data_emissao
#   col20 "STATUS"                    → status
# (NÚMERO RECAP / REVISÃO RECAP ficam intencionalmente fora — códigos alternativos
#  geralmente vazios; mapeá-los sobrescreveria o código oficial N-1710.)
COLUMN_MAP: dict[str, str] = {
    # codigo_documento — N-1710 é o número oficial da LD de engenharia
    "numero n 1710": "codigo_documento",
    "codigo documento": "codigo_documento",
    "codigo do documento": "codigo_documento",
    "codigo": "codigo_documento",
    "documento": "codigo_documento",
    "nº documento": "codigo_documento",
    "n documento": "codigo_documento",
    "numero documento": "codigo_documento",
    "numero do documento": "codigo_documento",
    "doc": "codigo_documento",
    "ld": "codigo_documento",
    # titulo
    "titulo": "titulo",
    "titulo do documento": "titulo",
    "descricao": "titulo",
    "descricao do documento": "titulo",
    "nome": "titulo",
    "nome do documento": "titulo",
    # disciplina
    "disciplina": "disciplina",
    "disc": "disciplina",
    "area": "disciplina",
    "especialidade": "disciplina",
    # revisao
    "revisao": "revisao",
    "rev": "revisao",
    "revisao atual": "revisao",
    # status
    "status": "status",
    "situacao": "status",
    "status sigem": "status",
    "status do documento": "status",
    "workflow": "status",
    "status workflow": "status",
    # a4_equivalente
    "a4 equivalente": "a4_equivalente",
    "a4equivalente": "a4_equivalente",
    "a4": "a4_equivalente",
    "qtd a4": "a4_equivalente",
    "quantidade a4": "a4_equivalente",
    "formato a4": "a4_equivalente",
    "peso a4": "a4_equivalente",
    # data_prevista
    "data prevista": "data_prevista",
    "previsto": "data_prevista",
    "data prevista emissao": "data_prevista",
    "data prevista de emissao": "data_prevista",
    "prazo": "data_prevista",
    "data programada": "data_prevista",
    # data_emissao
    "data emissao": "data_emissao",
    "emissao": "data_emissao",
    "data de emissao": "data_emissao",
    "data efetiva de emissao": "data_emissao",
    "data real": "data_emissao",
    "emitido em": "data_emissao",
}

CAMPOS = (
    "codigo_documento", "titulo", "disciplina", "revisao", "status",
    "a4_equivalente", "data_prevista", "data_emissao",
)

_ACENTOS = str.maketrans("áàâãäéèêëíìîïóòôõöúùûüç", "aaaaaeeeeiiiiooooouuuuc")


def _norm_header(v) -> str:
    """Normaliza um cabeçalho: minúsculas, sem acento, espaços colapsados."""
    if v is None:
        return ""
    s = str(v).strip().lower().translate(_ACENTOS)
    s = re.sub(r"[.\-_/]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _parse_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_float(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        # tenta direto (formato "1.5")
        try:
            return float(str(v).strip())
        except ValueError:
            return 0.0


def _achar_linha_cabecalho(rows: list[tuple], max_scan: int = 15) -> int:
    """Encontra a linha de cabeçalho: a que mapeia mais colunas conhecidas.

    LDs costumam ter linhas de título/logo antes do cabeçalho real.
    """
    melhor_idx, melhor_score = 0, -1
    for idx in range(min(max_scan, len(rows))):
        score = sum(1 for c in rows[idx] if _norm_header(c) in COLUMN_MAP)
        if score > melhor_score:
            melhor_idx, melhor_score = idx, score
    return melhor_idx


def _melhor_aba(wb, max_scan: int = 15) -> tuple:
    """Escolhe a aba cujo cabeçalho mapeia mais colunas conhecidas.

    A LD real tem abas CAPA / LISTA DE DOC. / BASE DADOS — só a do meio tem
    o cabeçalho de documentos. Pontuamos cada aba e ficamos com a melhor.
    """
    melhor_ws, melhor_rows, melhor_idx, melhor_score = wb.active, [], 0, -1
    for ws in wb.worksheets:
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(row)
            if i >= max_scan + 50:   # cabeçalho + amostra suficiente p/ varrer
                break
        if not rows:
            continue
        idx = _achar_linha_cabecalho(rows, max_scan)
        score = sum(1 for c in rows[idx] if _norm_header(c) in COLUMN_MAP)
        if score > melhor_score:
            melhor_ws, melhor_idx, melhor_score = ws, idx, score
    return melhor_ws, melhor_idx, melhor_score


def parse_ld(file_like: IO[bytes], aba: Optional[str] = None) -> dict:
    """Lê a LD e devolve {'documentos': [...], 'colunas_detectadas': {...},
    'aba': str, 'linha_cabecalho': int, 'ignoradas': int}."""
    wb = openpyxl.load_workbook(file_like, data_only=True, read_only=True)
    ws = wb[aba] if (aba and aba in wb.sheetnames) else _melhor_aba(wb)[0]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"documentos": [], "colunas_detectadas": {}, "aba": ws.title,
                "linha_cabecalho": 0, "ignoradas": 0}

    h_idx = _achar_linha_cabecalho(rows)
    header = rows[h_idx]
    col_para_campo: dict[int, str] = {}
    colunas_detectadas: dict[str, str] = {}
    for i, cell in enumerate(header):
        campo = COLUMN_MAP.get(_norm_header(cell))
        if campo and campo not in colunas_detectadas:
            col_para_campo[i] = campo
            colunas_detectadas[campo] = str(cell).strip() if cell is not None else ""

    documentos: list[dict] = []
    ignoradas = 0
    for row in rows[h_idx + 1:]:
        reg = {c: None for c in CAMPOS}
        for i, campo in col_para_campo.items():
            if i < len(row):
                reg[campo] = row[i]
        codigo = str(reg["codigo_documento"] or "").strip()
        if not codigo:
            ignoradas += 1
            continue
        documentos.append({
            "codigo_documento": codigo,
            "titulo": str(reg["titulo"]).strip() if reg["titulo"] else None,
            "disciplina": str(reg["disciplina"]).strip() if reg["disciplina"] else None,
            "revisao": str(reg["revisao"]).strip() if reg["revisao"] not in (None, "") else None,
            "status": str(reg["status"]).strip() if reg["status"] else None,
            "a4_equivalente": _parse_float(reg["a4_equivalente"]),
            "data_prevista": _parse_date(reg["data_prevista"]),
            "data_emissao": _parse_date(reg["data_emissao"]),
        })

    wb.close()
    return {
        "documentos": documentos,
        "colunas_detectadas": colunas_detectadas,
        "aba": ws.title,
        "linha_cabecalho": h_idx + 1,   # 1-based para exibição
        "ignoradas": ignoradas,
    }
