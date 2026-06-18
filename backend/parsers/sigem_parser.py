"""Parser da Consulta Geral do SIGEM.

Le o Excel "Consulta Geral - NOTIFICA SIGEM" e devolve documentos com status,
datas e hierarquia. O layout real tem cabecalho na linha 5, mas a deteccao
varre as primeiras linhas para continuar robusta a exportacoes futuras.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from typing import IO, Optional

import openpyxl


COLUMN_MAP: dict[str, str] = {
    "documento": "codigo_documento",
    "codigo documento": "codigo_documento",
    "codigo do documento": "codigo_documento",
    "numero documento": "codigo_documento",
    "numero do documento": "codigo_documento",
    "revisao": "revisao",
    "rev": "revisao",
    "revisao do documento ult sist": "revisao",
    "status": "status",
    "situacao": "status",
    "status sigem": "status",
    "modificado em": "modificado_em",
    "modificado": "modificado_em",
    "data modificacao": "modificado_em",
    "data de modificacao": "modificado_em",
    "incluido em": "incluido_em",
    "incluido": "incluido_em",
    "incluido no sistema": "incluido_em",
    "data inclusao": "incluido_em",
    "data de inclusao": "incluido_em",
    "nivel 1": "nivel_1",
    "nivel 2": "nivel_2",
    "nivel 3": "nivel_3",
    "nivel 4": "nivel_4",
    "nivel 5": "nivel_5",
    "nivel 6": "nivel_6",
    "nivel 7": "nivel_7",
    "nivel 8": "nivel_8",
}

CAMPOS = (
    "codigo_documento", "revisao", "status", "modificado_em", "incluido_em",
    "nivel_1", "nivel_2", "nivel_3", "nivel_4", "nivel_5", "nivel_6", "nivel_7", "nivel_8",
)

_ACENTOS = str.maketrans("áàâãäéèêëíìîïóòôõöúùûüçÁÀÂÃÄÉÈÊËÍÌÎÏÓÒÔÕÖÚÙÛÜÇ",
                         "aaaaaeeeeiiiiooooouuuucAAAAAEEEEIIIIOOOOOUUUUC")


def _norm_header(v) -> str:
    if v is None:
        return ""
    s = str(v).strip().lower().translate(_ACENTOS)
    s = re.sub(r"[.\-_/()]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _parse_datetime(v) -> Optional[datetime]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime.combine(v, datetime.min.time())
    s = str(v).strip()
    for fmt in (
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y %H:%M:%S", "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _achar_linha_cabecalho(rows: list[tuple], max_scan: int = 25) -> int:
    melhor_idx, melhor_score = 0, -1
    for idx in range(min(max_scan, len(rows))):
        score = sum(1 for c in rows[idx] if _norm_header(c) in COLUMN_MAP)
        if score > melhor_score:
            melhor_idx, melhor_score = idx, score
    return melhor_idx


def _melhor_aba(wb, max_scan: int = 25):
    melhor_ws, melhor_idx, melhor_score = wb.active, 0, -1
    for ws in wb.worksheets:
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(row)
            if i >= max_scan + 25:
                break
        if not rows:
            continue
        idx = _achar_linha_cabecalho(rows, max_scan)
        score = sum(1 for c in rows[idx] if _norm_header(c) in COLUMN_MAP)
        if score > melhor_score:
            melhor_ws, melhor_idx, melhor_score = ws, idx, score
    return melhor_ws, melhor_idx, melhor_score


def parse_sigem(file_like: IO[bytes], aba: Optional[str] = None) -> dict:
    """Le SIGEM e retorna dict com documentos e metadados da deteccao."""
    wb = openpyxl.load_workbook(file_like, data_only=True, read_only=True)
    ws = wb[aba] if (aba and aba in wb.sheetnames) else _melhor_aba(wb)[0]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"documentos": [], "colunas_detectadas": {}, "aba": ws.title,
                "linha_cabecalho": 0, "ignoradas": 0}

    h_idx = _achar_linha_cabecalho(rows)
    header = rows[h_idx]
    col_para_campo: dict[int, str] = {}
    colunas_detectadas: dict[str, str] = {}
    for i, cell in enumerate(header):
        campo = COLUMN_MAP.get(_norm_header(cell))
        if campo and campo not in colunas_detectadas:
            col_para_campo[i] = campo
            colunas_detectadas[campo] = str(cell).strip() if cell is not None else ""

    documentos: list[dict] = []
    ignoradas = 0
    for row in rows[h_idx + 1:]:
        reg = {c: None for c in CAMPOS}
        for i, campo in col_para_campo.items():
            if i < len(row):
                reg[campo] = row[i]
        codigo = str(reg["codigo_documento"] or "").strip()
        if not codigo:
            ignoradas += 1
            continue
        documentos.append({
            "codigo_documento": codigo,
            "revisao": str(reg["revisao"]).strip() if reg["revisao"] not in (None, "") else None,
            "status": str(reg["status"]).strip() if reg["status"] else None,
            "modificado_em": _parse_datetime(reg["modificado_em"]),
            "incluido_em": _parse_datetime(reg["incluido_em"]),
            **{
                f"nivel_{i}": str(reg[f"nivel_{i}"]).strip()
                if reg[f"nivel_{i}"] not in (None, "") else None
                for i in range(1, 9)
            },
        })

    wb.close()
    return {
        "documentos": documentos,
        "colunas_detectadas": colunas_detectadas,
        "aba": ws.title,
        "linha_cabecalho": h_idx + 1,
        "ignoradas": ignoradas,
    }
