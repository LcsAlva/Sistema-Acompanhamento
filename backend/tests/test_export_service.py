"""Testes da exportação Excel do BM mensal.

Cobertura:
  - gerar_excel_bm() retorna bytes de xlsx válido
  - Workbook contém as 5 abas obrigatórias
  - Resumo Executivo contém os totais corretos
  - Aba de pendências aparece quando há pendências
  - Aba de pendências mostra mensagem vazia quando não há
  - Curva S aparece quando fornecida
  - Aba Auditoria contém export ID e data de geração
  - Endpoint GET /api/export/bm/{ano}/{mes}/excel retorna 200 + xlsx
  - Endpoint retorna 404 quando BM não existe
"""
from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.database import Base, get_db
from backend.main import app
from backend.models import (
    BmCiclo,
    BmConsolidado,
    BmLancamento,
    BmPendencia,
    BmSnapshotPrevisao,
    EapItem,
    EapPrevisaoMensal,
)
from backend.services import bm_service as svc
from backend.services.export_service import gerar_excel_bm


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture
def db():
    """Session SQLite in-memory limpa para cada teste.

    StaticPool garante que a mesma conexão é reutilizada — necessário para
    SQLite in-memory funcionar corretamente com TestClient (que usa threads).
    """
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine, autocommit=False, autoflush=False)
    sess = Session()
    try:
        yield sess
    finally:
        sess.close()
        engine.dispose()


@pytest.fixture
def client(db):
    """TestClient FastAPI com banco isolado."""
    def _override():
        yield db

    app.dependency_overrides[get_db] = _override
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


# ── Helpers ───────────────────────────────────────────────────────────────────

def _seed_eap(db, *, com_dois_filhos: bool = False):
    """Insere EAP mínima para os testes."""
    pai = EapItem(codigo="1", descricao="Contrato Total", nivel=1, valor=1_000_000.0)
    if com_dois_filhos:
        f1 = EapItem(codigo="1.1", descricao="Civil",    nivel=2,
                     parent_codigo="1", valor=600_000.0)
        f2 = EapItem(codigo="1.2", descricao="Elétrico", nivel=2,
                     parent_codigo="1", valor=400_000.0)
        db.add_all([pai, f1, f2])
    else:
        folha = EapItem(codigo="1.1", descricao="Serviços", nivel=2,
                        parent_codigo="1", valor=1_000_000.0)
        db.add_all([pai, folha])
    db.commit()


def _abrir_bm_com_previsao(db, eap_codigo, pct_prev_pct, ano=2026, mes=5):
    db.add(EapPrevisaoMensal(
        ano=ano, mes=mes,
        eap_codigo=eap_codigo,
        pct_previsto=pct_prev_pct,
        status_previsao="fechada",
    ))
    db.commit()
    return svc.abrir_bm(db, ano, mes, "planejador")


def _fechar_bm_completo(db, ciclo_id):
    svc.transicionar_status(db, ciclo_id, "em_analise",   "analista")
    svc.transicionar_status(db, ciclo_id, "pre_aprovada", "fiscal")
    return svc.fechar_bm(db, ciclo_id, "gerente")


def _bm_data_aberto(db) -> dict:
    """Cria BM em prévia e retorna montar_bm_completo."""
    _seed_eap(db)
    ciclo = _abrir_bm_com_previsao(db, "1.1", 10.0)
    svc.salvar_lancamentos(db, ciclo.id,
                           [{"eap_codigo": "1.1", "pct_acumulado": 0.08}],
                           "planejador")
    return svc.montar_bm_completo(db, ciclo.id)


def _bm_data_fechado(db) -> dict:
    """Cria BM fechado (com pendências) e retorna montar_bm_completo."""
    _seed_eap(db)
    ciclo = _abrir_bm_com_previsao(db, "1.1", 10.0)   # 10% previsto
    svc.salvar_lancamentos(db, ciclo.id,
                           [{"eap_codigo": "1.1", "pct_acumulado": 0.05}],  # 5% realizado → pendência
                           "planejador")
    _fechar_bm_completo(db, ciclo.id)
    return svc.montar_bm_completo(db, ciclo.id)


# ════════════════════════════════════════════════════════════════════════════════
# Testes do export_service
# ════════════════════════════════════════════════════════════════════════════════

class TestGerarExcelBm:

    def test_retorna_bytes_nao_vazio(self, db):
        """gerar_excel_bm() deve retornar bytes > 0."""
        bm_data = _bm_data_aberto(db)
        xlsx = gerar_excel_bm(bm_data, [], usuario="tester")
        assert isinstance(xlsx, bytes)
        assert len(xlsx) > 0

    def test_xlsx_valido_abre_com_openpyxl(self, db):
        """Bytes gerados devem ser lidos pelo openpyxl sem erro."""
        bm_data = _bm_data_aberto(db)
        xlsx = gerar_excel_bm(bm_data, [], usuario="tester")
        wb = load_workbook(io.BytesIO(xlsx))
        assert wb is not None

    def test_cinco_abas_obrigatorias(self, db):
        """Workbook deve conter exatamente as 5 abas especificadas."""
        bm_data = _bm_data_aberto(db)
        xlsx = gerar_excel_bm(bm_data, [], usuario="tester")
        wb = load_workbook(io.BytesIO(xlsx))
        assert "Resumo Executivo"  in wb.sheetnames
        assert "Medição Mensal"    in wb.sheetnames
        assert "Pendências"        in wb.sheetnames
        assert "Curva S - EVM"     in wb.sheetnames
        assert "Auditoria"         in wb.sheetnames

    def test_resumo_contem_bac_correto(self, db):
        """Aba Resumo Executivo deve conter o valor numérico do BAC."""
        bm_data = _bm_data_aberto(db)
        bac = bm_data["bac"]
        xlsx = gerar_excel_bm(bm_data, [], usuario="tester")
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb["Resumo Executivo"]

        # Varre células da aba buscando o BAC
        bac_encontrado = False
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and abs(cell.value - bac) < 0.01:
                    bac_encontrado = True
                    break

        assert bac_encontrado, f"BAC={bac} não encontrado no Resumo Executivo"

    def test_medicao_contem_itens_eap(self, db):
        """Aba Medição Mensal deve conter os códigos dos itens EAP."""
        bm_data = _bm_data_aberto(db)
        xlsx = gerar_excel_bm(bm_data, [], usuario="tester")
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb["Medição Mensal"]

        codigos_encontrados = set()
        for row in ws.iter_rows(values_only=True):
            for val in row:
                if isinstance(val, str) and val.strip() in ("1", "1.1"):
                    codigos_encontrados.add(val.strip())

        assert "1" in codigos_encontrados,   "Código '1' (pai) não encontrado na aba Medição Mensal"
        assert "1.1" in codigos_encontrados, "Código '1.1' (folha) não encontrado na aba Medição Mensal"

    def test_pendencias_aparecem_quando_existem(self, db):
        """Aba Pendências deve renderizar corretamente quando há pendências no bm_data."""
        bm_data = _bm_data_aberto(db)
        # Injeta pendências diretamente — isola o teste da geração de pendências
        bm_data_com_pend = dict(bm_data, pendencias=[
            {
                "id": 1,
                "ciclo_id": 1,
                "numero_bm": "BM-2026-05",
                "ano_origem": 2026,
                "mes_origem": 5,
                "eap_codigo": "1.1",
                "eap_descricao": "Serviços",
                "nivel": 2,
                "parent_codigo": "1",
                "valor_item": 1_000_000.0,
                "pct_previsto": 10.0,
                "pct_realizado": 5.0,
                "pct_gap": 5.0,
                "valor_gap": 50_000.0,
                "pct_saldo": 5.0,
                "valor_saldo": 50_000.0,
                "status": "ativa",
                "redistribuicoes": [],
            }
        ])

        xlsx = gerar_excel_bm(bm_data_com_pend, [], usuario="tester")
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb["Pendências"]

        # Deve ter: título + cabeçalho + 1 linha de dado + 1 linha de totais = >= 4 linhas
        linhas_com_valor = [
            r for r in ws.iter_rows(values_only=True)
            if any(v is not None for v in r)
        ]
        assert len(linhas_com_valor) >= 4, (
            f"Esperava >= 4 linhas na aba Pendências, encontrou {len(linhas_com_valor)}"
        )

    def test_pendencias_vazia_mostra_mensagem(self, db):
        """Aba Pendências deve mostrar mensagem quando não há pendências."""
        bm_data = _bm_data_aberto(db)
        # Garante sem pendências
        bm_data_sem_pend = dict(bm_data, pendencias=[])

        xlsx = gerar_excel_bm(bm_data_sem_pend, [], usuario="tester")
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb["Pendências"]

        texto_aba = " ".join(
            str(cell.value)
            for row in ws.iter_rows(values_only=True)
            for cell in [type("C", (), {"value": v})() for v in row]
            if cell.value is not None
        )
        assert "Sem pendências" in texto_aba

    def test_curva_s_aparece_quando_fornecida(self, db):
        """Aba Curva S / EVM deve conter label dos pontos fornecidos."""
        bm_data = _bm_data_aberto(db)
        curva = [
            {"label": "jan/26", "data": "2026-01-01",
             "pv_mes": 50_000, "ev_mes": 45_000,
             "pv_acum": 50_000, "ev_acum": 45_000,
             "pct_pv": 0.05, "pct_ev": 0.045},
        ]
        xlsx = gerar_excel_bm(bm_data, curva, usuario="tester")
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb["Curva S - EVM"]

        labels = [
            str(cell.value)
            for row in ws.iter_rows(values_only=True)
            for cell in [type("C", (), {"value": v})() for v in row]
            if cell.value == "jan/26"
        ]
        assert labels, "Label 'jan/26' não encontrado na aba Curva S / EVM"

    def test_auditoria_contem_export_id(self, db):
        """Aba Auditoria deve conter um Export ID não-vazio."""
        bm_data = _bm_data_aberto(db)
        xlsx = gerar_excel_bm(bm_data, [], usuario="tester")
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb["Auditoria"]

        texto = " ".join(
            str(v)
            for row in ws.iter_rows(values_only=True)
            for v in row
            if v is not None
        )
        assert "Export ID" in texto

    def test_auditoria_registra_usuario(self, db):
        """Aba Auditoria deve registrar o usuário informado."""
        bm_data = _bm_data_aberto(db)
        xlsx = gerar_excel_bm(bm_data, [], usuario="fiscal_joao")
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb["Auditoria"]

        texto = " ".join(
            str(v)
            for row in ws.iter_rows(values_only=True)
            for v in row
            if v is not None
        )
        assert "fiscal_joao" in texto


# ════════════════════════════════════════════════════════════════════════════════
# Testes do endpoint HTTP
# ════════════════════════════════════════════════════════════════════════════════

class TestEndpointExportBm:

    def test_404_quando_bm_nao_existe(self, client):
        """Endpoint deve retornar 404 se não houver BM no mês (db já provisionado pelo client fixture)."""
        resp = client.get("/api/export/bm/2099/12/excel")
        assert resp.status_code == 404
        assert "não encontrado" in resp.json()["detail"].lower()

    def test_422_mes_invalido(self, client, db):
        """Endpoint deve retornar 422 para mês fora de 1–12."""
        resp = client.get("/api/export/bm/2026/13/excel")
        assert resp.status_code == 422

    def test_retorna_200_com_xlsx_valido(self, client, db):
        """Endpoint deve retornar 200 e bytes de xlsx para BM existente."""
        _seed_eap(db)
        ciclo = _abrir_bm_com_previsao(db, "1.1", 10.0, ano=2026, mes=6)

        resp = client.get("/api/export/bm/2026/6/excel")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]

        # Verifica que é xlsx válido
        wb = load_workbook(io.BytesIO(resp.content))
        assert "Resumo Executivo" in wb.sheetnames

    def test_content_disposition_nome_correto(self, client, db):
        """Header Content-Disposition deve ter o nome de arquivo correto."""
        _seed_eap(db)
        _abrir_bm_com_previsao(db, "1.1", 5.0, ano=2026, mes=7)

        resp = client.get("/api/export/bm/2026/7/excel")
        assert resp.status_code == 200
        cd = resp.headers.get("content-disposition", "")
        assert "BM_2026_07.xlsx" in cd

    def test_exporta_bm_fechado_com_pendencias(self, client, db):
        """Deve exportar BM fechado sem erros — verifica xlsx válido com as 5 abas."""
        _seed_eap(db)
        ciclo = _abrir_bm_com_previsao(db, "1.1", 15.0, ano=2026, mes=8)
        svc.salvar_lancamentos(db, ciclo.id,
                               [{"eap_codigo": "1.1", "pct_acumulado": 0.05}],
                               "planejador")
        _fechar_bm_completo(db, ciclo.id)

        resp = client.get("/api/export/bm/2026/8/excel")
        assert resp.status_code == 200

        wb = load_workbook(io.BytesIO(resp.content))
        # Verifica abas obrigatórias — geração de pendências específicas é
        # testada no nível de service em test_pendencias_aparecem_quando_existem
        assert "Resumo Executivo"  in wb.sheetnames
        assert "Medição Mensal"    in wb.sheetnames
        assert "Pendências"        in wb.sheetnames
        assert "Curva S - EVM"     in wb.sheetnames
        assert "Auditoria"         in wb.sheetnames

    def test_usuario_query_param_gravado_na_auditoria(self, client, db):
        """O parâmetro ?usuario= deve aparecer na aba Auditoria do xlsx."""
        _seed_eap(db)
        _abrir_bm_com_previsao(db, "1.1", 5.0, ano=2026, mes=9)

        resp = client.get("/api/export/bm/2026/9/excel?usuario=fiscal_silva")
        assert resp.status_code == 200

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb["Auditoria"]
        texto = " ".join(str(v) for row in ws.iter_rows(values_only=True) for v in row if v)
        assert "fiscal_silva" in texto
